"""N5 系 ACIC 控制器配置字 Excel 解析器。

本地 Excel 格式（"N5系控制器配置字" sheet）:
  - Row 0-2: 标题行（车型配置表、车型代号等）
  - Row 3: 列标题 (DID, 适配零件号, 字节, 位, 配置项名称, 值, 值描述, ...)
  - Row 4+: 数据行
    - 配置项定义行: 字节(col2)有数字, 位(col3)有 "BitX-BitY(dec)" 格式
    - 值选项行: 配置项下面的多行，仅 值(col5)/值描述(col6) 有内容
    - 分隔行: 字节列有 "ByteN（hex）" 格式
    - 总值行、填写说明行: 跳过

输出: list[ConfigItem] — 标准化的配置项列表
"""

import re
import openpyxl
from lib.parsers import ConfigItem, register_parser

_BIT_RANGE_PAT = re.compile(r'Bit(\d+)-Bit(\d+)')
_BIT_SINGLE_PAT = re.compile(r'Bit(\d+)')

HEADER_ROW = 3  # 0-indexed row that contains column headers
DATA_START_ROW = 4


def _parse_bit_count(bit_str):
    """Parse 'Bit7-Bit6(dec)' → 2, or 'Bit5(dec)' → 1."""
    m = _BIT_RANGE_PAT.search(bit_str)
    if m:
        return int(m.group(1)) - int(m.group(2)) + 1
    m = _BIT_SINGLE_PAT.search(bit_str)
    if m:
        return 1
    return 0


def _clean_name(raw_name):
    """Strip \\nreserved / \\nresversed suffixes and whitespace."""
    name = raw_name.replace('\r\n', '\n')
    for suffix in ('\nreserved', '\nresversed'):
        if name.lower().endswith(suffix):
            name = name[:len(name) - len(suffix)]
    return name.replace('\n', '').strip()


def _is_reserved(raw_name):
    """Determine if a config item is truly reserved.
    Items like 'AEB自动紧急制动系统\\nreserved' are NOT reserved —
    only items with no real name (empty, '预留', bare 'reserved') are."""
    cleaned = _clean_name(raw_name)
    if not cleaned:
        return True
    lower = cleaned.lower()
    return lower in ('reserved', 'resversed', '预留')


def _cell_str(val):
    return str(val).strip() if val is not None else ''


@register_parser("n5_baic_acic")
@register_parser("n5x_acic")
def parse(excel_path, sheet_name):
    """Parse N5系 ACIC controller config Excel into ConfigItem list."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    items = []
    current_byte = None

    for i in range(DATA_START_ROW, len(rows)):
        row = rows[i]
        if len(row) < 5:
            continue

        byte_val = _cell_str(row[2])
        bit_str = _cell_str(row[3])
        name_raw = _cell_str(row[4])
        value_desc = _cell_str(row[6]) if len(row) > 6 else ""

        if byte_val and byte_val.isdigit():
            current_byte = int(byte_val)

        # Skip separator rows like "Byte0（hex）", summary rows, instructions
        if 'Byte' in byte_val and '（' in byte_val:
            continue
        if byte_val == '总值' or byte_val.startswith('填写说明'):
            continue

        # Only process rows that define a config item (have Bit info)
        if not bit_str or 'Bit' not in bit_str:
            continue

        bit_count = _parse_bit_count(bit_str)
        if bit_count == 0:
            continue

        clean = _clean_name(name_raw) if name_raw else ''
        reserved = _is_reserved(name_raw) if name_raw else True

        if current_byte is None:
            continue

        # For items with no name (pure reserved), give a placeholder
        if not clean:
            clean = ''
            reserved = True

        items.append(ConfigItem(
            byte=current_byte,
            bit_count=bit_count,
            type="uint8_t",
            chinese_name=clean,
            description=value_desc,
            is_reserved=reserved,
            raw_chinese_name=name_raw,
        ))

    return items
