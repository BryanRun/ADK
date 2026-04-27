"""捷途 T1V「03.3.Coding DID」表解析器。

仅处理 DID **0xF011**（整车配置信息）块：从首次出现 0xF011 的行起向下扫描。

列索引（0-based）: 6=Byte, 7=Bit, 8=英文名, 9=中文名
"""

import re
import openpyxl
from lib.parsers import ConfigItem, register_parser

_NOT_USED = re.compile(r"not\s*used", re.I)


def _cell_str(val):
    return str(val).strip() if val is not None else ""


def _to_macro_name(name: str) -> str:
    if not name:
        return ""
    name = re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_").upper()
    if name == "VEHICLE_TYPE":
        name = "VEHICLE_TYPE_MODE"
    if name and name[0].isdigit():
        name = f"N_{name}"
    return name


def _clean_name(s):
    if not s:
        return ""
    return _cell_str(s).replace("\r\n", "\n").replace("\n", "")


def _is_reserved_cn_en(cn, en):
    c, e = _clean_name(cn), _clean_name(en)
    if not c and not e:
        return True
    if _NOT_USED.search(c) or _NOT_USED.search(e):
        return True
    if c in ("预留", "/") or e.lower() in ("reserved",):
        return True
    return False


def _bit_width(bit_cell):
    if bit_cell is None:
        return None
    if isinstance(bit_cell, bool):
        return 1
    if isinstance(bit_cell, (int, float)):
        return 1
    s = str(bit_cell).strip()
    if s.upper() == "ALL":
        return 8
    for sep in ("~", "-"):
        if sep in s:
            a, b = [x.strip() for x in s.split(sep, 1)]
            return int(b) - int(float(a)) + 1
    try:
        float(s)
        return 1
    except ValueError:
        return 8


def _byte_list(byte_cell, prev_byte):
    """Return list of byte indices for this row; may be empty."""
    if byte_cell is None:
        return [prev_byte] if prev_byte is not None else []
    if isinstance(byte_cell, bool):
        return [int(byte_cell)]
    if isinstance(byte_cell, (int, float)):
        return [int(byte_cell)]
    s = str(byte_cell).strip()
    for sep in ("~", "-"):
        if sep in s and not s.lower().startswith("0x"):
            a, b = [x.strip() for x in s.split(sep, 1)]
            lo, hi = int(float(a)), int(float(b))
            return list(range(lo, hi + 1))
    try:
        return [int(float(s))]
    except ValueError:
        return []


def _split_width_across_bytes(byte_list, total_width, bit_cell_str):
    """Emit (byte, bit_count) pairs.

    total_width 是 _bit_width 返回的单字节位宽（如 0~7 → 8）。
    多字节范围时，每个字节分配相同的位宽。
    """
    if not byte_list:
        return []
    w = min(total_width, 8)
    return [(b, w) for b in byte_list]


@register_parser("jetour_t1v_coding")
def parse(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    start = None
    for i, row in enumerate(rows):
        if row and len(row) > 1 and _cell_str(row[1]) == "0xF011":
            start = i
            break
    if start is None:
        raise ValueError("jetour_t1v_coding: 未找到 DID 0xF011 行")

    items = []
    prev_byte = None
    bit_cell_str = ""

    for row in rows[start:]:
        if not row or len(row) < 10:
            continue
        bcell = row[6] if len(row) > 6 else None
        bit_cell = row[7] if len(row) > 7 else None
        en = row[8] if len(row) > 8 else None
        cn = row[9] if len(row) > 9 else None
        extra_desc = _cell_str(row[13]) if len(row) > 13 else ""

        blist = _byte_list(bcell, prev_byte)
        if not blist:
            continue

        prev_byte = blist[-1]
        w = _bit_width(bit_cell)
        if w is None or w <= 0:
            continue

        bit_cell_str = str(bit_cell).strip() if bit_cell is not None else ""
        pairs = _split_width_across_bytes(blist, w, bit_cell_str)

        cn_s = _clean_name(cn)
        en_s = _clean_name(en)
        display_name = cn_s or en_s
        reserved = _is_reserved_cn_en(cn, en)

        cn_raw = _cell_str(cn)
        val_desc = _cell_str(row[13]) if len(row) > 13 else ""
        raw_line = "\n".join(x for x in (cn_raw, val_desc) if x)

        for bnum, bc in pairs:
            items.append(
                ConfigItem(
                    byte=bnum,
                    bit_count=min(bc, 8),
                    type="uint8_t",
                    chinese_name=display_name if not reserved else "",
                    english_name=_to_macro_name(en_s) if (en_s and not reserved) else "",
                    description=extra_desc,
                    is_reserved=reserved,
                    raw_chinese_name=raw_line,
                )
            )

    if not items:
        raise ValueError("jetour_t1v_coding: 0xF011 块未解析到任何配置项")

    return items
