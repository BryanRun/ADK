"""北汽 N80 ICC 配置码表解析器。

工作表「ICC配置码表」列约定:
  col0 DID, col1 位置(Bit0~Bit7 / ByteN（hex）), col2 配置项名称, col3 物理值定义

ByteN（hex）行出现在每一字节的位定义之后；其后 Bit 行属于字节 (N+1)。
首组 Bit（在第一个 Byte 分隔符之前）属于字节 0。
"""

import re
import openpyxl
from lib.parsers import ConfigItem, register_parser

_BYTE_MARK = re.compile(r"^Byte(\d+)（hex）$")
_BIT_MARK = re.compile(r"^Bit(\d+)$", re.I)


def _cell_str(val):
    return str(val).strip() if val is not None else ""


def _clean_name(raw_name):
    name = raw_name.replace("\r\n", "\n") if raw_name else ""
    for suffix in ("\nreserved", "\nresversed"):
        if name.lower().endswith(suffix):
            name = name[: len(name) - len(suffix)]
    return " ".join(name.replace("\n", " ").split())


def _is_reserved(raw_name):
    cleaned = _clean_name(raw_name)
    if not cleaned:
        return True
    lower = cleaned.lower()
    return lower in ("reserved", "resversed", "预留")


@register_parser("baic_n80_icc")
def parse(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    items = []
    current_byte = 0
    seen_any_bit = False

    for row in rows[6:]:
        if len(row) < 3:
            continue
        pos = _cell_str(row[1])
        name_raw = _cell_str(row[2]) if len(row) > 2 else ""
        phys_desc = _cell_str(row[3]) if len(row) > 3 else ""

        m_byte = _BYTE_MARK.match(pos)
        if m_byte:
            n = int(m_byte.group(1))
            current_byte = n + 1
            continue

        m_bit = _BIT_MARK.match(pos)
        if not m_bit:
            continue

        seen_any_bit = True
        clean = _clean_name(name_raw) if name_raw else ""
        reserved = _is_reserved(name_raw) if name_raw else True
        if not clean:
            reserved = True

        items.append(
            ConfigItem(
                byte=current_byte,
                bit_count=1,
                type="uint8_t",
                chinese_name=clean,
                description=phys_desc,
                is_reserved=reserved,
                raw_chinese_name=name_raw,
            )
        )

    if not seen_any_bit:
        raise ValueError("baic_n80_icc: 未解析到任何 Bit 行，请检查表头与 sheet 名")

    return items
