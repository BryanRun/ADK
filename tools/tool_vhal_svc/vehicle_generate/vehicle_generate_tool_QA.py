# coding=UTF-8
import copy
import os
from re import T
import re
import sys
import warnings
import json
from openpyxl import load_workbook
from collections import defaultdict

# Common Part
MAP_TITLE_ROW = 2
MAP_COTENT_ROW = 3
g_title_name_list = 'PropID,PropID Value,AreaID,size(Byte),initValue,AreaType,ValueType,Access,ChangeMode,Function Description,' \
                    'minValue,maxValue,OS_Target,PropID_Android,ValueType_Android,AreaType_Android,AreaID_VHal,Value Description'
                    
g_title_name_svw_list = 'PropID,PropID Value,AreaID,size(Byte),initValue,AreaType,ValueType,Access,ChangeMode,Function Description,' \
                    'minValue,maxValue,OS_Target,PropID,ValueType,AreaType,AreaID_VHal,Value Description'

#g_exclude_list_names = ['MFS_INCFOLWDST_514','MFS_DECFOLWDST_514','MFS_CUSTBTN_514','MFS_PARKAID_514','MFS_LEROLLUP_514','MFS_LEROLLDWN_514', \
#                        'MFS_LEROLLPRESS_514','MFS_SRCSWTBTN_514','MFS_VOICERCTCNBTN_514','MFS_PREVSONGTUNESIG_514','MFS_NEXTSONGTUNESIG_514', \
#                        'MFS_RIROLLUP_514','MFS_RIROLLDWN_514','MFS_RIROLLPRESS_514']
g_exclude_list_names = []

g_title_col_dict = {}
g_viclient_signal_list = ''
g_initPropValue_list = ''
g_initPropConfig_list = ''
g_android_subscribe_list = ''
g_qnx_subscribe_list = ''
g_android_propConfig_list = ''
g_getPropValueID_list = ''
g_types_hal_list = ''
g_vehicle_property_Ids_list = ''
g_android_write_int32vec_list = ''
g_initPropJson_list = []
g_initPropCfgJson_list = []

g_need_creat_Signal_H = True #True False
g_need_creat_Signal_Proto = True #True False
g_need_creat_InitVehiclePropValue_CPP = False #True False
g_need_creat_InitVehiclePropConfig_CPP = False #True False
g_need_creat_SubscribeProps_CPP = True #True False
g_need_creat_QNX_Subscribe_New_CPP = False #True False
g_need_creat_GetPropValueID_CPP = True #True False
g_need_creat_InitPropConfigs_CPP = True #True False
g_need_creat_Types_Hal_New = True #True False
g_need_creat_ALVehiclePropertyIds_JAVA = True #True False
g_need_creat_VehicleHal_setVehiclePropValue_CPP = False #True False
g_need_creat_vehicle_prop_value_json = True #True False
g_need_creat_vehicle_prop_config_json = True #True False

isDebug = False #True False

##########
## 生成 vehicle_prop_value.json 文件 数据
## 生成VehicleIF缓存PropValue json文件
##
##
## 全局变量 g_initPropJson_list
##
##########
def generate_initPropValue_json(sheet):
    global g_initPropJson_list

    # 去除重复READ的propID value，保持表中顺序
    propId_Value_col = g_title_col_dict["PropID Value"]
    access_col = g_title_col_dict['Access']
    propId_Value_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        access = sheet.cell(row, access_col).value
        if access == "READ":
            propId_value = sheet.cell(row, propId_Value_col).value
            propId_Value_list.append(propId_value)
    propId_Value_list = remove_same_prop_elements(propId_Value_list)
    print('[Information] initPropValue_json((Access_Read) total num: ' + str(len(propId_Value_list)))

    # 每个PropID有多少条Read Signal
    propid_col = g_title_col_dict['PropID']
    mark_dict = {}
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        prop_value = sheet.cell(row, propId_Value_col).value
        access = sheet.cell(row, access_col).value
        if access == "READ":
            if prop_value not in mark_dict.keys():
                mark_dict[prop_value] = 1
            else:
                mark_dict[prop_value] = mark_dict[prop_value] + 1
    # print(mark_dict)
    prop_tag_mark = copy.deepcopy(mark_dict) # 用来表示X, ViClientPropValue propVal_X;

    # 按去重后prop列表的顺序生成代码
    label = 0
    valueType_col = g_title_col_dict['ValueType']
    areaId_col = g_title_col_dict['AreaID']
    byteSize_col = g_title_col_dict['size(Byte)']
    initValue_col = g_title_col_dict['initValue']
    for propId_Value in propId_Value_list:
        label = label + 1
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            prop_dict = {}
            propid = sheet.cell(row, propid_col).value
            prop_value = sheet.cell(row, propId_Value_col).value
            access = sheet.cell(row, access_col).value
            valueType = sheet.cell(row, valueType_col).value
            areaId = sheet.cell(row, areaId_col).value
            byteSize = sheet.cell(row, byteSize_col).value
            if(byteSize is None or byteSize == ''):
                byteSize = 0
            initValue = sheet.cell(row, initValue_col).value
            if (prop_value == propId_Value) and (access == "READ"):
                if mark_dict[prop_value] > 1:
                    subTag = prop_tag_mark[prop_value]
                    if mark_dict[prop_value] == subTag:
                        prop_tag_mark[prop_value] = prop_tag_mark[prop_value] - 1
                    else:
                        continue
                prop_dict["propId"] = int(prop_value, 16)
                prop_dict["valueType"] = handle_value_type(valueType)
                prop_dict["areaId"] = 0
                prop_dict["initialValue"] = int(initValue, 16)
                prop_dict["size"] = int(byteSize / 4)
                g_initPropJson_list.append(prop_dict)


# 生成VehicleIF缓存PropConfig

##########
## 生成 vehicle_prop_config.json 文件 数据
## 生成VehicleIF PropConfig json文件
##
##
## 全局变量 g_initPropCfgJson_list
##
##########
def generate_initPropConfigJson_file(sheet):
    global g_initPropCfgJson_list

    # 去除重复READ的propID value，保持表中顺序
    propId_Value_col = g_title_col_dict["PropID Value"]
    propId_Value_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        propId_value = sheet.cell(row, propId_Value_col).value
        propId_Value_list.append(propId_value)
    propId_Value_list = remove_same_prop_elements(propId_Value_list)
    print('[Information] initPropConfigJson total num: ' + str(len(propId_Value_list)))

    # 查找异常Case
    target_col = g_title_col_dict['OS_Target']
    access_col = g_title_col_dict['Access']
    areaID_col = g_title_col_dict['AreaID']
    # {propID : [target1, target2,...] , }
    prop_os_dict = {}
    # 查找同一PropID下，一部分是OS_QNX，另一部分是OS_Android
    same_prop_same_area_diff_target_list = []
    same_prop_diff_area_diff_target_list = []
    for propId_Value in propId_Value_list:
        frist_areaID = ''
        prop_os_dict[propId_Value] = []
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            prop_value = sheet.cell(row, propId_Value_col).value
            target = sheet.cell(row, target_col).value
            areaID = sheet.cell(row, areaID_col).value
            if target == '':
                target = 'OS_ALL'
            if propId_Value == prop_value:
                if len(prop_os_dict[propId_Value]) == 0:
                    prop_os_dict[propId_Value].append(target)
                    frist_areaID = areaID
                else:
                    if target not in prop_os_dict[propId_Value]:
                        prop_os_dict[propId_Value].append(target)
                        if areaID == frist_areaID:
                            same_prop_same_area_diff_target_list.append(propId_Value)
                        else:
                            same_prop_diff_area_diff_target_list.append(propId_Value)
    if len(same_prop_same_area_diff_target_list) != 0:
        print('[Information] same propId same areaId has diff os_target: ', same_prop_same_area_diff_target_list)
        # print(same_prop_same_area_diff_target_list)
    if len(same_prop_diff_area_diff_target_list) != 0:
        print('[Information] same propId diff areaId has diff os_target: ', same_prop_diff_area_diff_target_list)
        # print(same_prop_diff_area_diff_target_list)

    #  建立prop-area-access之间的字典
    # {propid1: {areaid1:access, areaid2:access, ...}, propid2: {areaid1:access, areaid2:access, ...}}
    prop_area_access_dict = {}
    for propId_Value in propId_Value_list:
        prop_area_access_dict[propId_Value] = {}
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            prop_value = sheet.cell(row, propId_Value_col).value
            areaID = sheet.cell(row, areaID_col).value
            access = sheet.cell(row, access_col).value
            if prop_value == propId_Value:
                if areaID in prop_area_access_dict[propId_Value].keys():
                    # 如果存在该areaID，比较access
                    if prop_area_access_dict[propId_Value][areaID] != access:
                        prop_area_access_dict[propId_Value][areaID] = "READ_WRITE"
                    else:
                        print("[info] same prop same area has same access: " + propId_Value)
                else:
                    prop_area_access_dict[propId_Value][areaID] = access

    # 建立PropID_Access
    # 同一个PropID，对应一个Access, 如果有READ_WRITE, 则认为是READ_WRITE。如果存在一个area是读，一个area是写，单独处理。
    # 个人建议：access放入AreaConfig中，如果需要就修改
    prop_access_dict = {}
    for prop, innerDict in prop_area_access_dict.items():
        access_list = innerDict.values()
        if 'READ_WRITE' in access_list:
            access = 'READ_WRITE'
        else:
            if ('READ' in access_list) and ('WRITE' in access_list):
                access = 'READ_WRITE'
            else:
                # access列表中只有一个元素
                if len(list(set(access_list))) != 1:
                    print('[error] temp list: ' + access_list)
                access = list(access_list)[0]
        prop_access_dict[prop] = access

    # Config共同部分  {propID_Value: [propID, areaType, valueType, changeMode]}
    # 默认一个PropID下这些信息是一致的，即便不一致，也处理为一致
    propid_common_info_dict = {}
    changeMode_col = g_title_col_dict['ChangeMode']
    valueType_col = g_title_col_dict['ValueType']
    propId_col = g_title_col_dict['PropID']
    areaType_col = g_title_col_dict['AreaType']
    propId_Value_col = g_title_col_dict["PropID Value"]

    for propId_Value in propId_Value_list:
        propid_common_info_dict[propId_Value] = []
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            prop_value = sheet.cell(row, propId_Value_col).value
            if prop_value == propId_Value:
                propId = sheet.cell(row, propId_col).value
                changeMode = sheet.cell(row, changeMode_col).value
                valueType = sheet.cell(row, valueType_col).value
                areaType = sheet.cell(row, areaType_col).value
                propValue = sheet.cell(row, propId_Value_col).value
                propid_common_info_dict[propId_Value].append(propValue)
                propid_common_info_dict[propId_Value].append(areaType)
                propid_common_info_dict[propId_Value].append(valueType)
                propid_common_info_dict[propId_Value].append(changeMode)
    #            propid_common_info_dict[propId_Value].append(propValue)

    # 生成PropID_AreaID_MAX_MIN, 默认相同PropID和AreaID拥有相同的max和min值
    # {propID_Value: {areaID : [min, max]}}
    maxValue_col = g_title_col_dict['maxValue']
    minValue_col = g_title_col_dict['minValue']
    prop_access_min_max_dict = {}
    for propId_Value in propId_Value_list:
        prop_access_min_max_dict[propId_Value] = {}
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            prop = sheet.cell(row, propId_Value_col).value
            area = sheet.cell(row, areaID_col).value
            min = sheet.cell(row, minValue_col).value
            max = sheet.cell(row, maxValue_col).value
            if propId_Value == prop:
                if area not in prop_access_min_max_dict[propId_Value].keys():
                    prop_access_min_max_dict[propId_Value][area] = [min, max]
                else:
                    min_max = prop_access_min_max_dict[propId_Value][area]
                    if min_max[0] > min:
                        min_max[0] = min
                    if min_max[1] < max:
                        min_max[1] = max
                    prop_access_min_max_dict[propId_Value][area] = min_max



    # 生成主题代码
    # 相同PropID，不同OS—Target的特殊处理
    label = 0
    for prop, innerDict in prop_area_access_dict.items():
        prop_dict = {}
        prop_dict["propId"] = int(propid_common_info_dict[prop][0], 16)
        prop_dict["areaType"] = handle_area_type(propid_common_info_dict[prop][1])
        prop_dict["valueType"] = handle_value_type(propid_common_info_dict[prop][2])
        prop_dict["access"] = handle_access(prop_access_dict[prop])
        prop_dict["changeMode"] = handle_change_mode(propid_common_info_dict[prop][3])
        area_list = []

        label = label + 1
        sublabel = len(innerDict)
        for area, access in innerDict.items():
            area_dict = {}
            minValue = prop_access_min_max_dict[prop][area][0]
            maxValue = prop_access_min_max_dict[prop][area][1]
            area_dict["areaId"] = 0
            area_dict["minValue"] = int(minValue, 16)
            if (maxValue == 'INT32_MAX') or (maxValue == '0xFFFFFFFF'):
                area_dict["maxValue"] = 2147483647
            else:
                area_dict["maxValue"] = int(maxValue, 16)
            area_list.append(area_dict)
        prop_dict["areaCfg"] = area_list
        for os_target in prop_os_dict[prop]:
            prop_dict["osTarget"] = handle_os_target(os_target)
        g_initPropCfgJson_list.append(prop_dict)

def handle_value_type(s: str) -> int:
    mapping = {
        "STRING": 0,
        "BOOL": 1,
        "INT32": 2,
        "INT32_VEC": 3,
        "INT64": 4,
        "INT64_VEC": 5,
        "FLOAT": 6,
        "FLOAT_VEC": 7,
        "BYTES": 8,
    }
    return mapping.get(s.upper(), 2)

def handle_area_type(s: str) -> int:
    mapping = {
        "GLOBAL": 1,
        "WINDOW": 3,
        "MIRROR": 4,
        "SEAT": 5,
        "DOOR": 6,
        "WHEEL": 7,
        "BODY": 8,
    }
    return mapping.get(s.upper(), 1)

def handle_access(s: str) -> int:
    mapping = {
        "NONE": 0,
        "READ": 1,
        "WRITE": 2,
        "READ_WRITE": 3,
    }
    return mapping.get(s.upper(), 1)

def handle_change_mode(s: str) -> int:
    mapping = {
        "STATIC": 0,
        "ON_CHANGE": 1,
        "CONTINOUS": 2,
    }
    return mapping.get(s.upper(),1)

def handle_os_target(s: str) -> int:
    mapping = {
        "INVALID": 0,
        "OS_QNX": 1,
        "OS_ANDROID": 2,
        "OS_ALL": 3,
        "MASK": 0xFFFF,
    }
    return mapping.get(s.upper(),0)

##########
## 生成ViClient中Siganal_{project}头文件 数据
## Map sheet 【PropID】 [PropID Value]队列函数
##
## 需要用到表格 [PropID],[PropID Value]列
## 全局变量 g_viclient_signal_list
##
## 	VCU_GEARSIGVLD_214                                           =0x1000,
##	VCU_GEARSIG_214                                              =0x1001,
##########
def generate_viclient_signal_code(sheet):
    global g_viclient_signal_list

    propId_col = g_title_col_dict["PropID"]
    propId_Value_col = g_title_col_dict["PropID Value"]
    dict = {} # 用来存储PropID 和 PropID_Value
    special_dict = {} # 用来存储相同PropID，不同Prop_Value

    # 判断是否存在相同PropId，不同PropName
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        propId = sheet.cell(row, propId_col).value
        propId_value = sheet.cell(row, propId_Value_col).value
        if propId_value.startswith('0x'):
            if propId_value not in dict.keys():
                dict[propId_value] = propId
                if (isDebug):
                    print(f"PropID {propId}, {propId_value}")
            elif dict[propId_value] != propId:
                #  如果该PropID Value之前添加进去，则判断PropId是否相同
                special_dict[propId_value] = propId
                print("[error] same propId diff propId value: ", propId_value)
        else:
            print('[error] PropID Value error: ' + propId_value)


    # 生成Signal_*.h 主体代码
    count = 0
    for PropId, PropName in dict.items():
        g_viclient_signal_list += '\t%-60s =%s,\n' % (PropName, PropId)
        count = count + 1
    for PropId, PropName in special_dict.items():
        g_viclient_signal_list += '\t%-60s =%s,\n' % (PropName, PropId)
        count = count + 1
    print("[Information] ViClient Signal.h PropID total num: " + str(count))

##########
## 生成InitVehiclePropValue_New.cpp文件 数据
## 生成VehicleIF缓存PropValue
##
## 全局变量 g_initPropValue_list
##
##
##########
def generate_initPropValue_code(sheet):
    global g_initPropValue_list

    # 去除重复READ的propID value，保持表中顺序
    propId_Value_col = g_title_col_dict["PropID Value"]
    access_col = g_title_col_dict['Access']
    propId_Value_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        access = sheet.cell(row, access_col).value
        if access == "READ":
            propId_value = sheet.cell(row, propId_Value_col).value
            propId_Value_list.append(propId_value)

    if(isDebug):
        print(f"prop_elements len: {len(propId_Value_list)}")

    propId_Value_list = remove_same_prop_elements(propId_Value_list)
    print('[Information] PropID_Value(Access_Read) total num: ' + str(len(propId_Value_list)))

    # 每个PropID有多少条Read Signal
    propid_col = g_title_col_dict['PropID']
    mark_dict = {}
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        prop_value = sheet.cell(row, propId_Value_col).value
        access = sheet.cell(row, access_col).value
        if access == "READ":
            if prop_value not in mark_dict.keys():
                mark_dict[prop_value] = 1
            else:
                mark_dict[prop_value] = mark_dict[prop_value] + 1
    #if(isDebug):
        #print(mark_dict)
    prop_tag_mark = copy.deepcopy(mark_dict) # 用来表示X, ViClientPropValue propVal_X;

    # 按去重后prop列表的顺序生成代码
    label = 0
    valueType_col = g_title_col_dict['ValueType']
    areaId_col = g_title_col_dict['AreaID']
    byteSize_col = g_title_col_dict['size(Byte)']
    initValue_col = g_title_col_dict['initValue']
    for propId_Value in propId_Value_list:
        label = label + 1
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            propid = sheet.cell(row, propid_col).value
            prop_value = sheet.cell(row, propId_Value_col).value
            access = sheet.cell(row, access_col).value
            valueType = sheet.cell(row, valueType_col).value
            areaId = sheet.cell(row, areaId_col).value
            byteSize = sheet.cell(row, byteSize_col).value
            initValue = sheet.cell(row, initValue_col).value
            if (prop_value == propId_Value) and (access == "READ"):#?????为什么只记录只读的
                # 判断该PropID是否只有一条可读信号
                if mark_dict[prop_value] == 1:
                    # 如果只有一条信号
                    g_initPropValue_list += '\t// PRO_%s\n' % propid
                    g_initPropValue_list += '\tViClientPropValue propVal%d;\n' % label
                    g_initPropValue_list += '\tRawValue* valPtr%d =  propVal%d.mutable_data();\n' % (label, label)
                    g_initPropValue_list += '\tpropVal%d.set_propid(PRO_%s);\n' % (label, propid)
                    g_initPropValue_list += '\tpropVal%d.set_valuetype(autolink::ValueType::%s);\n' % (label, valueType)
                    g_initPropValue_list += '\tpropVal%d.set_areaid(AreaID_%s);\n' % (label, areaId)
                    if valueType == "INT32_VEC": #数组
                        try:
                            g_initPropValue_list += '\tfor(int i = 0; i<%d; i++)\n' % (byteSize/4)
                        except:
                            g_initPropValue_list += '\tfor(int i = 0; i<%d; i++)\n' % (int(byteSize) / 4)
                        g_initPropValue_list += '\t\tvalPtr%d->add_int32values(0x%x);\n' % (label, int(initValue, 16))
                    elif valueType == "INT32":
                        g_initPropValue_list += '\tvalPtr%d->add_int32values(0x%x);\n' % (label, int(initValue, 16))
                    g_initPropValue_list += '\tVehicleProperty vehicleProp%d;\n' % label
                    g_initPropValue_list += '\tvehicleProp%d.initPropVector(propVal%d);\n' % (label, label)
                    g_initPropValue_list += '\tm_Properties.insert(std::map<int32_t, VehicleProperty>::value_type(PRO_%s, vehicleProp%d));\n' % (propid, label)
                    g_initPropValue_list += '\n'
                elif mark_dict[prop_value] > 1:
                    # 如果同一PropId包含多条Signal
                    subTag = prop_tag_mark[prop_value]
                    if mark_dict[prop_value] == subTag:
                        # 如果是第一次出现该PropID
                        g_initPropValue_list += '\t// PRO_%s\n' % propid
                        g_initPropValue_list += '\tViClientPropValue propVal%d_%d;\n' % (label, subTag)
                        g_initPropValue_list += '\tRawValue* valPtr%d_%d =  propVal%d_%d.mutable_data();\n' % (label, subTag, label, subTag)
                        g_initPropValue_list += '\tpropVal%d_%d.set_propid(PRO_%s);\n' % (label, subTag, propid)
                        g_initPropValue_list += '\tpropVal%d_%d.set_valuetype(autolink::ValueType::%s);\n' % (label, subTag, valueType)
                        g_initPropValue_list += '\tpropVal%d_%d.set_areaid(AreaID_%s);\n' % (label, subTag, areaId)
                        if valueType == "INT32_VEC":
                            g_initPropValue_list += '\tfor(int i = 0; i<%d; i++)\n' % (byteSize / 4)
                            g_initPropValue_list += '\t\tvalPtr%d_%d->add_int32values(0x%x);\n' % (label, subTag, int(initValue, 16))
                        elif valueType == "INT32":
                            g_initPropValue_list += '\tvalPtr%d_%d->add_int32values(0x%x);\n' % (label, subTag, int(initValue, 16))
                        g_initPropValue_list += '\tVehicleProperty vehicleProp%d;\n' % label
                        g_initPropValue_list += '\tvehicleProp%d.initPropVector(propVal%d_%d);\n' % (label, label, subTag)
                        g_initPropValue_list += '\n'
                        prop_tag_mark[prop_value] = prop_tag_mark[prop_value] - 1
                    elif subTag < mark_dict[prop_value]:
                        # 如果不是第一次出现, 可能是最后一次
                        g_initPropValue_list += '\tViClientPropValue propVal%d_%d;\n' % (label, subTag)
                        g_initPropValue_list += '\tRawValue* valPtr%d_%d =  propVal%d_%d.mutable_data();\n' % (label, subTag, label, subTag)
                        g_initPropValue_list += '\tpropVal%d_%d.set_propid(PRO_%s);\n' % (label, subTag, propid)
                        g_initPropValue_list += '\tpropVal%d_%d.set_valuetype(autolink::ValueType::%s);\n' % (label, subTag, valueType)
                        g_initPropValue_list += '\tpropVal%d_%d.set_areaid(AreaID_%s);\n' % (label, subTag, areaId)
                        if valueType == "INT32_VEC":
                            g_initPropValue_list += '\tfor(int i = 0; i<%d; i++)\n' % (byteSize / 4)
                            g_initPropValue_list += '\t\tvalPtr%d_%d->add_int32values(0x%x);\n' % (label, subTag, int(initValue, 16))
                        elif valueType == "INT32":
                            g_initPropValue_list += '\tvalPtr%d_%d->add_int32values(0x%x);\n' % (label, subTag, int(initValue, 16))
                        g_initPropValue_list += '\tvehicleProp%d.initPropVector(propVal%d_%d);\n' % (label, label, subTag)
                        prop_tag_mark[prop_value] = prop_tag_mark[prop_value] - 1
                        if prop_tag_mark[prop_value] == 0:
                            g_initPropValue_list += '\tm_Properties.insert(std::map<int32_t, VehicleProperty>::value_type(PRO_%s, vehicleProp%d));\n' % (propid, label)
                        g_initPropValue_list += '\n'

##########
## 生成InitVehiclePropConfig_New.cpp文件 数据
## 生成VehicleIF缓存PropConfig
##
## 全局变量 g_initPropConfig_list
##
##
##########
def generate_initPropConfig_code(sheet):
    global g_initPropConfig_list

    # 去除重复READ的propID value，保持表中顺序
    propId_Value_col = g_title_col_dict["PropID Value"]
    propId_Value_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        propId_value = sheet.cell(row, propId_Value_col).value
        propId_Value_list.append(propId_value)
    propId_Value_list = remove_same_prop_elements(propId_Value_list)
    # print('[Information] remove same READ+WRITE propid value after total num: ' + str(len(propId_Value_list)))

    # 查找异常Case
    target_col = g_title_col_dict['OS_Target']
    access_col = g_title_col_dict['Access']
    areaID_col = g_title_col_dict['AreaID']
    # {propID : [target1, target2,...] , }
    prop_os_dict = {}
    # 查找同一PropID下，一部分是OS_QNX，另一部分是OS_Android
    same_prop_same_area_diff_target_list = []
    same_prop_diff_area_diff_target_list = []
    for propId_Value in propId_Value_list:
        frist_areaID = ''
        prop_os_dict[propId_Value] = []
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            prop_value = sheet.cell(row, propId_Value_col).value
            target = sheet.cell(row, target_col).value
            areaID = sheet.cell(row, areaID_col).value
            if target == '':
                target = 'OS_ALL'
            if propId_Value == prop_value:
                if len(prop_os_dict[propId_Value]) == 0:
                    prop_os_dict[propId_Value].append(target)
                    frist_areaID = areaID
                else:
                    if target not in prop_os_dict[propId_Value]:
                        prop_os_dict[propId_Value].append(target)
                        if areaID == frist_areaID:
                            same_prop_same_area_diff_target_list.append(propId_Value)
                        else:
                            same_prop_diff_area_diff_target_list.append(propId_Value)
    if len(same_prop_same_area_diff_target_list) != 0:
        print('[Information] same propId same areaId has diff os_target: ', same_prop_same_area_diff_target_list)
        # print(same_prop_same_area_diff_target_list)
    if len(same_prop_diff_area_diff_target_list) != 0:
        print('[Information] same propId diff areaId has diff os_target: ', same_prop_diff_area_diff_target_list)
        # print(same_prop_diff_area_diff_target_list)

    #  建立prop-area-access之间的字典
    # {propid1: {areaid1:access, areaid2:access, ...}, propid2: {areaid1:access, areaid2:access, ...}}
    prop_area_access_dict = {}

    for propId_Value in propId_Value_list:
        prop_area_access_dict[propId_Value] = {}
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            prop_value = sheet.cell(row, propId_Value_col).value
            areaID = sheet.cell(row, areaID_col).value
            access = sheet.cell(row, access_col).value
            if prop_value == propId_Value:
                if areaID in prop_area_access_dict[propId_Value].keys():
                    # 如果存在该areaID，比较access
                    if prop_area_access_dict[propId_Value][areaID] != access:
                        prop_area_access_dict[propId_Value][areaID] = "READ_WRITE"
                    else:
                        print("[info] same prop same area has same access: " + propId_Value)
                else:
                    prop_area_access_dict[propId_Value][areaID] = access

    # 建立PropID_Access
    # 同一个PropID，对应一个Access, 如果有READ_WRITE, 则认为是READ_WRITE。如果存在一个area是读，一个area是写，单独处理。
    # 个人建议：access放入AreaConfig中，如果需要就修改
    prop_access_dict = {}
    for prop, innerDict in prop_area_access_dict.items():
        access_list = innerDict.values()
        if 'READ_WRITE' in access_list:
            access = 'READ_WRITE'
        else:
            if ('READ' in access_list) and ('WRITE' in access_list):
                access = 'READ_WRITE'
            else:
                # access列表中只有一个元素
                if len(list(set(access_list))) != 1:
                    print('[error] temp list: ' + access_list)
                access = list(access_list)[0]
        prop_access_dict[prop] = access

    # Config共同部分  {propID_Value: [propID, areaType, valueType, changeMode]}
    # 默认一个PropID下这些信息是一致的，即便不一致，也处理为一致
    propid_common_info_dict = {}
    changeMode_col = g_title_col_dict['ChangeMode']
    valueType_col = g_title_col_dict['ValueType']
    propId_col = g_title_col_dict['PropID']
    areaType_col = g_title_col_dict['AreaType']

    for propId_Value in propId_Value_list:
        propid_common_info_dict[propId_Value] = []
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            prop_value = sheet.cell(row, propId_Value_col).value
            if prop_value == propId_Value:
                propId = sheet.cell(row, propId_col).value
                changeMode = sheet.cell(row, changeMode_col).value
                valueType = sheet.cell(row, valueType_col).value
                areaType = sheet.cell(row, areaType_col).value
                propid_common_info_dict[propId_Value].append(propId)
                propid_common_info_dict[propId_Value].append(areaType)
                propid_common_info_dict[propId_Value].append(valueType)
                propid_common_info_dict[propId_Value].append(changeMode)

    # 生成PropID_AreaID_MAX_MIN, 默认相同PropID和AreaID拥有相同的max和min值
    # {propID_Value: {areaID : [min, max]}}
    maxValue_col = g_title_col_dict['maxValue']
    minValue_col = g_title_col_dict['minValue']
    prop_access_min_max_dict = {}
    for propId_Value in propId_Value_list:
        prop_access_min_max_dict[propId_Value] = {}
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            prop = sheet.cell(row, propId_Value_col).value
            area = sheet.cell(row, areaID_col).value
            min = sheet.cell(row, minValue_col).value
            max = sheet.cell(row, maxValue_col).value
            if propId_Value == prop:
                if area not in prop_access_min_max_dict[propId_Value].keys():
                    prop_access_min_max_dict[propId_Value][area] = [int(min, 16), int(max,16)]


    # 生成主题代码
    # 相同PropID，不同OS—Target的特殊处理
    label = 0
    for prop, innerDict in prop_area_access_dict.items():
        label = label + 1
        g_initPropConfig_list += '\t// PRO_%s\n' % propid_common_info_dict[prop][0]
        g_initPropConfig_list += '\tViClientPropConfig propConfig%d;\n' % label
        g_initPropConfig_list += '\tpropConfig%d.set_propid(PRO_%s);\n' % (label, propid_common_info_dict[prop][0])
        g_initPropConfig_list += '\tpropConfig%d.set_areatype(AreaType::%s);\n' % (label, propid_common_info_dict[prop][1])
        g_initPropConfig_list += '\tpropConfig%d.set_valuetype(ValueType::%s);\n' % (label, propid_common_info_dict[prop][2])
        g_initPropConfig_list += '\tpropConfig%d.set_access(Access::%s);\n' % (label, prop_access_dict[prop])
        g_initPropConfig_list += '\tpropConfig%d.set_changemode(ChangeMode::%s);\n' % (label, propid_common_info_dict[prop][3])
        sublabel = len(innerDict)
        for area, access in innerDict.items():
            minValue = prop_access_min_max_dict[prop][area][0]
            maxValue = prop_access_min_max_dict[prop][area][1]
            if len(innerDict) == 1:
                # 如果该PropID下，只有一个AreaID
                g_initPropConfig_list += '\tAreaConfig * areaConfig%d = propConfig%d.add_areaconfigs();\n' % (label, label)
                g_initPropConfig_list += '\tareaConfig%d->set_areaid(AreaID_%s);\n' % (label, area)
                if maxValue == 'INT32_MAX':
                    g_initPropConfig_list += '\tareaConfig%d->set_maxint32value(U%s);\n' % (label, maxValue)
                else:
                    g_initPropConfig_list += '\tareaConfig%d->set_maxint32value(%d);\n' % (label, int(maxValue))
                g_initPropConfig_list += '\tareaConfig%d->set_minint32value(%d);\n' % (label, int(minValue))
            else:
                # 如果该PropID下，有多个areaID
                g_initPropConfig_list += '\tAreaConfig * areaConfig%d_%d = propConfig%d.add_areaconfigs();\n' % (label, sublabel, label)
                g_initPropConfig_list += '\tareaConfig%d_%d->set_areaid(AreaID_%s);\n' % (label, sublabel, area)
                if maxValue == 'INT32_MAX':
                    g_initPropConfig_list += '\tareaConfig%d_%d->set_maxint32value(U%s);\n' % (label, sublabel, maxValue)
                else:
                    g_initPropConfig_list += '\tareaConfig%d_%d->set_maxint32value(%d);\n' % (label, sublabel, int(maxValue))
                g_initPropConfig_list += '\tareaConfig%d_%d->set_minint32value(%d);\n' % (label, sublabel, int(minValue))
                sublabel = sublabel - 1
        # 当前PropID相同, OS_Target不同的, 都是areaID相同，一个可读给OS_ALL，一个可写OS_Android
        # 如果一个PropID相同, areaID不同，OS-Target不同，access不同情况下，需要分开处理。当前没有处理，问题不大，即便GetConfig来判断是否可写，VehicleIF也不会Set
        for os_target in prop_os_dict[prop]:
            if os_target == 'OS_ALL':
                g_initPropConfig_list += '\tconfigs_ALL.add_config()->CopyFrom(propConfig%d);\n' % label
            elif os_target == 'OS_QNX':
                g_initPropConfig_list += '\tconfigs_QNX.add_config()->CopyFrom(propConfig%d);\n' % label
            elif os_target == 'OS_ANDROID':
                g_initPropConfig_list += '\tconfigs_Android.add_config()->CopyFrom(propConfig%d);\n' % label
        g_initPropConfig_list += '\n'

##########
## 生成 SubscribeProps.cpp 文件 数据
## 生成Android客户端订阅代码
##
## 全局变量 g_android_subscribe_list
##
##
##########
def generate_android_subscribe_code(sheet, projectName):
    global g_android_subscribe_list
    # 去除重复READ的propID value，保持表中顺序
    PropID_Android_col = g_title_col_dict["PropID_Android"]
    access_col = g_title_col_dict['Access']
    target_col = g_title_col_dict['OS_Target']
    PropID_col = g_title_col_dict['PropID']
    # try:
    #     project_col = g_title_col_dict['427']
    # except:
    #     print('更改项目名project_col = g_title_col_dict[project]')
    # project_col = 1
    propId_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        access = sheet.cell(row, access_col).value
        target = sheet.cell(row, target_col).value
        PropID_Android = sheet.cell(row, PropID_Android_col).value
        PropID = sheet.cell(row, PropID_col).value
        # try:
        #     project = sheet.cell(row, project_col).value
        # except:
        #     project = ''
        # 存在部分信号access == "READ"  target == "" PropID_Android != ''
        if access == "READ" and target != 'OS_QNX' and PropID_Android != '':
            propId_list.append(PropID_Android)
        # 检测异常情况, 避免QNX PropID 转换 Android PropID时出错
        if access == "READ" and (target == 'OS_ALL' or target == 'OS_ANDROID') and PropID_Android == '':
            # if 'SIMU+' in project:
            print('[error] PropID = %s row = %d PropID_Android is null ' % (PropID, row))
    propId_list = remove_same_prop_elements(propId_list)
    print('[Information] Subscribe Android PropID(Access_Read) total num: ' + str(len(propId_list)))

    # 生成订阅代码
    g_android_subscribe_list += '        int32_t subscribeList[] = {\n'
    for propID in propId_list:
        g_android_subscribe_list += '\t\t\ttoInt(%s::%s),\n' % (projectName, propID)
    g_android_subscribe_list += '        };\n'
    # print(g_subscribe_list)

##########
## 生成 QNX_Subscribe_New.cpp 文件 数据
## 生成QNX客户端订阅代码
##
## 全局变量 g_qnx_subscribe_list
##
##
##########
def generate_qnx_subscribe_code(sheet, project):
    global g_qnx_subscribe_list
    # 去除重复READ的propID value，保持表中顺序
    PropID_col = g_title_col_dict["PropID"]
    access_col = g_title_col_dict['Access']
    target_col = g_title_col_dict['OS_Target']
    propId_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        access = sheet.cell(row, access_col).value
        target = sheet.cell(row, target_col).value
        PropID = sheet.cell(row, PropID_col).value
        # 存在部分信号access == "READ"  target == "" PropID_Android != ''
        if access == "READ" and target != 'OS_ANDROID' and PropID != '':
            propId_list.append(PropID)
    propId_list = remove_same_prop_elements(propId_list)
    print('[Information] Subscribe QNX PropID(Access_Read) total num: ' + str(len(propId_list)))

    for propID in propId_list:
        g_qnx_subscribe_list += '\ttoInt(%s::%s),\n' % (project, propID)


##########
## 生成 GetPropValueID.cpp 文件 数据
## 生成VehicleHal中VehicleHalFunc::getPropValueID方法
## getPropValueID是将QNX侧广播的Signal的PropID转为Android PropID
##
## 全局变量 g_getPropValueID_list
##
##
##########
def generate_getPropValueID_code(sheet, project):
    global g_getPropValueID_list

    # 去除重复READ的propID value，保持表中顺序
    PropID_Android_col = g_title_col_dict["PropID_Android"]
    propID_col = g_title_col_dict["PropID"]
    access_col = g_title_col_dict['Access']
    target_col = g_title_col_dict['OS_Target']
    propId_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        access = sheet.cell(row, access_col).value
        target = sheet.cell(row, target_col).value
        PropID_Android = sheet.cell(row, PropID_Android_col).value
        propID = sheet.cell(row, propID_col).value
        # 存在部分信号access == "READ"  target == "" PropID_Android != ''
        if access == "READ" and target != 'OS_QNX' and PropID_Android != '':
            if PropID_Android == propID:
                propId_list.append(PropID_Android)
            else:
                print('[error] Android PropID = %s diff with QNX PropID = %s' % (PropID_Android, propID))
    propId_list = remove_same_prop_elements(propId_list)
    print('[Information] Android getPropValueID PropID(Access_Read) total num: ' + str(len(propId_list)))

    # 生成代码
    g_getPropValueID_list += '\t%s pid = static_cast<%s>(propId);\n' % (project, project)
    g_getPropValueID_list += '\t\tswitch(pid) {\n'
    for propID in propId_list:
        g_getPropValueID_list += '\t\t\tcase %s::%s:\n' % (project, propID)
        g_getPropValueID_list += '\t\t\t{\n'
        g_getPropValueID_list += '\t\t\t\treturn VendorVehicleProperty::%s;\n' % propID
        g_getPropValueID_list += '\t\t\t}\n'
    g_getPropValueID_list += '\t\t\tdefault:\n'
    g_getPropValueID_list += '\t\t\t{\n'
    g_getPropValueID_list += '\t\t\t\tALOGD("[%s] propId = 0x%04x out of range", __func__, propId);\n'
    g_getPropValueID_list += '\t\t\t\treturn VendorVehicleProperty::VENDOR_PROPERTY_INVALID;\n'
    g_getPropValueID_list += '\t\t\t}\n'
    g_getPropValueID_list += '\t\t}\n'

##########
## 生成 InitPropConfigs.cpp 文件 数据
## 生成PropertyConfigManager代码
##
## 全局变量 g_android_propConfig_list
##
##
##########
def generate_android_propConfig_code(sheet):
    global g_android_propConfig_list

    # 去除重复PropID_Android，保持表中顺序
    PropID_Android_col = g_title_col_dict["PropID_Android"]
    access_col = g_title_col_dict['Access']
    AreaID_Vhal_col = g_title_col_dict['AreaID_VHal']
    ValueType_Android_col = g_title_col_dict["ValueType_Android"]
    initValue_col = g_title_col_dict['initValue']
    propId_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        PropID_Android = sheet.cell(row, PropID_Android_col).value
        # 存在部分信号access == "READ"  target == "" PropID_Android != ''
        if PropID_Android.startswith('PDU_TPMS'):#过滤胎压，单独处理，手写在template 表
            continue
        if PropID_Android != '':
            propId_list.append(PropID_Android)
    propId_list = remove_same_prop_elements(propId_list)
    print('[Information] Android propConfig PropID total num: ' + str(len(propId_list)))

    # 建立PropID_Android_access_changeMode关系
    PropID_Access_dict = {}
    PropID_ChangeMode_dict = {}
    PropID_AreaID_Vhal_dict = {}
    PropID_initValue_dict = {}
    PropID_ValueType_dict = {}
    changeMode_col = g_title_col_dict['ChangeMode']
    for propID in propId_list:
        PropID_Access_dict[propID] = []
        PropID_ChangeMode_dict[propID] = []
        PropID_AreaID_Vhal_dict[propID] = []
        PropID_initValue_dict[propID] = []
        PropID_ValueType_dict[propID] = []
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            PropID_Android = sheet.cell(row, PropID_Android_col).value
            access = sheet.cell(row, access_col).value
            changeMode = sheet.cell(row, changeMode_col).value
            AreaID_Vhal = sheet.cell(row, AreaID_Vhal_col).value
            initvalue = sheet.cell(row, initValue_col).value
            valueType = sheet.cell(row, ValueType_Android_col).value
            if(initvalue is None or initvalue == ''):
                print('[error] propID = %s initvalue is null' % PropID_Android)
            if propID == PropID_Android:
                PropID_Access_dict[propID].append(access)
                PropID_ChangeMode_dict[propID].append(changeMode)
                if(valueType == 'STRING'):
                    PropID_ValueType_dict[propID].append('stringValue')
                    PropID_initValue_dict[propID].append(initvalue)
                elif (valueType == 'INT32'):
                    PropID_ValueType_dict[propID].append('int32Values')
                    PropID_initValue_dict[propID].append(int(initvalue, 16))
                elif (valueType == 'INT32_VEC'):
                    PropID_ValueType_dict[propID].append('int32Values')
                    PropID_initValue_dict[propID].append(initvalue)
                elif (valueType == 'INT64'):
                    PropID_ValueType_dict[propID].append('int64Values')
                    PropID_initValue_dict[propID].append(int(initvalue, 16))
                elif (valueType == 'INT64_VEC'):
                    PropID_ValueType_dict[propID].append('int64Values')
                    PropID_initValue_dict[propID].append(initvalue)
                elif (valueType == 'FLOAT'):
                    PropID_ValueType_dict[propID].append('floatValues')
                    PropID_initValue_dict[propID].append(initvalue)
                elif (valueType == 'FLOAT_VEC'):
                    PropID_ValueType_dict[propID].append('floatValues')
                    PropID_initValue_dict[propID].append(initvalue)
                elif (valueType == 'BYTES'):
                    PropID_ValueType_dict[propID].append('bytes')
                    PropID_initValue_dict[propID].append(initvalue)
                if AreaID_Vhal not in PropID_AreaID_Vhal_dict[propID]:
                    PropID_AreaID_Vhal_dict[propID].append(AreaID_Vhal)

    # 合并READ_WRITE 和 找到特殊ChangeMode
    for propID in propId_list:
        # if {'WRITE', 'READ'}.issubset(PropID_Access_dict[propID]):
        if 'WRITE' in PropID_Access_dict[propID] and 'READ' in PropID_Access_dict[propID]:
            PropID_Access_dict[propID].clear()
            PropID_Access_dict[propID].append('READ_WRITE')
        else:
            if 'WRITE' in PropID_Access_dict[propID]:
                PropID_Access_dict[propID].clear()
                PropID_Access_dict[propID].append('WRITE')
            elif 'READ' in PropID_Access_dict[propID]:
                PropID_Access_dict[propID].clear()
                PropID_Access_dict[propID].append('READ')
            else:
                print('[error] propID = %s Access is null' % propID)

        if {'STATIC', 'Trigger', 'CONTINOUS'}.issubset(PropID_ChangeMode_dict[propID]):
            print('[warning] exist special ChangeMode: ' + propID)
        else:
            PropID_ChangeMode_dict[propID].clear()
            PropID_ChangeMode_dict[propID].append('ON_CHANGE')

    # 生成主体代码
    # g_android_propConfig_list += 'PropertyConfigManager::PropertyConfigManager()\n'
    # g_android_propConfig_list += ': m_propConfigs({\n'
    #g_android_propConfig_list += 'm_propConfigs = {\n'
    #g_android_propConfig_list += '\t\t\t{\n'
    #g_android_propConfig_list += '\t\t\t\t.prop = static_cast<int32_t>(VehicleProperty::VHAL_HEARTBEAT),\n'
    #g_android_propConfig_list += '\t\t\t\t.access = VehiclePropertyAccess::READ,\n'
    #g_android_propConfig_list += '\t\t\t\t.changeMode = VehiclePropertyChangeMode::ON_CHANGE,\n'
    #g_android_propConfig_list += '\t\t\t\t.areaConfigs = {{.areaId = 0x0000}},\n'
    #g_android_propConfig_list += '\t\t\t},\n'
    for propID in propId_list:
        g_android_propConfig_list += '\t\t\t{.config ={\n'
        g_android_propConfig_list += '\t\t\t\t.prop = static_cast<int32_t>(VendorVehicleProperty::%s),\n' % propID
        g_android_propConfig_list += '\t\t\t\t.access = VehiclePropertyAccess::%s,\n' % PropID_Access_dict[propID][0]
        g_android_propConfig_list += '\t\t\t\t.changeMode = VehiclePropertyChangeMode::%s,\n' % PropID_ChangeMode_dict[propID][0]
        if len(PropID_AreaID_Vhal_dict[propID]) == 1:
            g_android_propConfig_list += '\t\t\t\t.areaConfigs = {{.areaId = %s}},\n' % PropID_AreaID_Vhal_dict[propID][0]
        else:
            g_android_propConfig_list += '\t\t\t.areaConfigs = {'
            cout = len(PropID_AreaID_Vhal_dict[propID])
            for areaid_vahl in PropID_AreaID_Vhal_dict[propID]:
                if cout > 1:
                    g_android_propConfig_list += '{.areaId = %s}, ' % areaid_vahl
                else:
                    g_android_propConfig_list += '{.areaId = %s}' % areaid_vahl
                cout = cout - 1
            g_android_propConfig_list += '\t},\n'
        g_android_propConfig_list += '\t\t\t},\n'
        g_android_propConfig_list += '\t\t\t.initialValue = {.%s' % PropID_ValueType_dict[propID][0]
        if(PropID_ValueType_dict[propID][0] == 'floatValues'):
            g_android_propConfig_list += ' = {%sf}}},\n' % PropID_initValue_dict[propID][0]
        else:
            g_android_propConfig_list += ' = {%s}}},\n' % PropID_initValue_dict[propID][0]
    #g_android_propConfig_list += '\t\t};\n'
    # g_android_propConfig_list += '})\n'
    # g_android_propConfig_list += '{}\n'

##########
## 生成 Types_Hal_New.cpp 文件 数据
## 生成自定义types.hal接口
## 对于相同propID，不同PropName的，按两信号处理
##
## 全局变量 g_types_hal_list
##
##
##########
def generate_types_hal_code(sheet):
    global g_types_hal_list

    Function_Description_col = g_title_col_dict["Function Description"]
    Value_Description_col = g_title_col_dict["Value Description"]
    PropID_Android_col = g_title_col_dict["PropID_Android"]
    ValueType_Android_col = g_title_col_dict["ValueType_Android"]
    PropID_Value_col = g_title_col_dict["PropID Value"]
    AreaType_Android_col = g_title_col_dict["AreaType_Android"]
    ChangeMode_col = g_title_col_dict['ChangeMode']
    Access_col = g_title_col_dict['Access']

    # 去除重复的PropID_Android
    propId_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        PropID_Android = sheet.cell(row, PropID_Android_col).value
        if PropID_Android != '':
            propId_list.append(PropID_Android)
    propId_list = remove_same_prop_elements(propId_list)
    print('[Information] Android Types.hal PropID total num: ' + str(len(propId_list)))

    #  建立PropID_Value-ValueType-AreaType
    PropID_AreaType_dict = {}
    for propID in propId_list:
        PropID_AreaType_dict[propID] = {'PropID_Value':'', 'ValueType':'', 'AreaType':'', 'ChangeMode':'', 'Access':'', 'v_Description':'', 'f_Description':''}
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            PropID_Android = sheet.cell(row, PropID_Android_col).value
            PropID_Value = sheet.cell(row, PropID_Value_col).value
            ValueType = sheet.cell(row, ValueType_Android_col).value
            AreaType = sheet.cell(row, AreaType_Android_col).value
            ChangeMode = sheet.cell(row, ChangeMode_col).value
            Access = sheet.cell(row, Access_col).value
            v_Description = sheet.cell(row, Value_Description_col).value
            f_Description = sheet.cell(row, Function_Description_col).value
            if propID == PropID_Android:
                if PropID_AreaType_dict[PropID_Android]['PropID_Value'] == '':
                    PropID_AreaType_dict[PropID_Android]['PropID_Value'] = PropID_Value
                elif PropID_AreaType_dict[PropID_Android]['PropID_Value'] != PropID_Value:
                    print('[error] same PropID_Android has diff PropID_Value: ' + PropID_Android)
                if PropID_AreaType_dict[PropID_Android]['ValueType'] == '':
                    PropID_AreaType_dict[PropID_Android]['ValueType'] = ValueType
                elif PropID_AreaType_dict[PropID_Android]['ValueType'] != ValueType:
                    print('[error] same PropID_Android has diff ValueType: ' + PropID_Android)
                if PropID_AreaType_dict[PropID_Android]['AreaType'] == '':
                    PropID_AreaType_dict[PropID_Android]['AreaType'] = AreaType
                elif PropID_AreaType_dict[PropID_Android]['AreaType'] != AreaType:
                    print('[error] same PropID_Android has diff AreaType: ' + PropID_Android)
                if PropID_AreaType_dict[PropID_Android]['ChangeMode'] == '':
                    PropID_AreaType_dict[PropID_Android]['ChangeMode'] = ChangeMode
                elif PropID_AreaType_dict[PropID_Android]['ChangeMode'] != ChangeMode:
                    print('[error] same PropID_Android has diff ChangeMode: ' + PropID_Android)
                if PropID_AreaType_dict[PropID_Android]['Access'] == '':
                    PropID_AreaType_dict[PropID_Android]['Access'] = Access
                elif PropID_AreaType_dict[PropID_Android]['Access'] != Access:
                    PropID_AreaType_dict[PropID_Android]['Access'] = 'READ_WRITE'
                if (v_Description == '' or v_Description is None):
                    PropID_AreaType_dict[PropID_Android]['v_Description'] = ''
                else:
                    Description = v_Description.replace('\n', "\n\t\t")
                    Description = '\n\t\t'+Description.strip()
                    PropID_AreaType_dict[PropID_Android]['v_Description'] = Description
                if (f_Description == '' or f_Description is None):
                    PropID_AreaType_dict[PropID_Android]['f_Description'] = ''
                else:
                    PropID_AreaType_dict[PropID_Android]['f_Description'] = f_Description

    # 生成主体代码
    for PropID_Android in propId_list:
        PropID_Value = PropID_AreaType_dict[PropID_Android]['PropID_Value']
        ValueType = PropID_AreaType_dict[PropID_Android]['ValueType']
        AreaType = PropID_AreaType_dict[PropID_Android]['AreaType']
        ChangeMode = PropID_AreaType_dict[PropID_Android]['ChangeMode']
        Access = PropID_AreaType_dict[PropID_Android]['Access']
        v_Description = PropID_AreaType_dict[PropID_Android]['v_Description']
        f_Description = PropID_AreaType_dict[PropID_Android]['f_Description']
        g_types_hal_list +='''
    /**
     * %s
     *
     * Function Describe: %s
     * VehiclePropertyChangeMode:%s
     * VehiclePropertyAccess:%s
     * Signal Value:%s
     */
    ''' % (PropID_Android, f_Description, ChangeMode, Access, v_Description)
        g_types_hal_list += '%s = (\n' % PropID_Android
        g_types_hal_list += '        %s\n' % PropID_Value
        g_types_hal_list += '        | VehiclePropertyGroup:VENDOR\n'
        g_types_hal_list += '        | VehiclePropertyType:%s\n' % ValueType
        if AreaType == 'BODY':
            g_types_hal_list += '        | VendorVehicleArea:%s),' % AreaType
        else:
            g_types_hal_list += '        | VehicleArea:%s),' % AreaType
        g_types_hal_list += '\n'
        #print(g_types_hal_list)

##########
## 生成 ALVehiclePropertyIds.java 文件 数据
## 生成自定义ALVehiclePropertyIds接口
## 对于相同propID，不同PropName的，按两信号处理
##
## 全局变量 g_vehicle_property_Ids_list
##
##
##########
def generate_ALVehiclePropertyIds_code(sheet):
    global g_vehicle_property_Ids_list

    Function_Description_col = g_title_col_dict["Function Description"]
    Value_Description_col = g_title_col_dict["Value Description"]
    PropID_Android_col = g_title_col_dict["PropID_Android"]
    ValueType_Android_col = g_title_col_dict["ValueType_Android"]
    PropID_Value_col = g_title_col_dict["PropID Value"]
    AreaType_Android_col = g_title_col_dict["AreaType_Android"]
    ChangeMode_col = g_title_col_dict['ChangeMode']
    Access_col = g_title_col_dict['Access']

    # 去除重复的PropID_Android
    propId_list = []
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        PropID_Android = sheet.cell(row, PropID_Android_col).value
        if PropID_Android in g_exclude_list_names:
            continue
        if PropID_Android != '':
            propId_list.append(PropID_Android)
    propId_list = remove_same_prop_elements(propId_list)
    print('[Information] Android Types.hal PropID total num: ' + str(len(propId_list)))

    #  建立PropID_Value-ValueType-AreaType
    PropID_AreaType_dict = {}
    for propID in propId_list:
        PropID_AreaType_dict[propID] = {'PropID_Value':'', 'ValueType':'', 'AreaType':'', 'ChangeMode':'', 'Access':'', 'v_Description':'', 'f_Description':''}
        for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
            PropID_Android = sheet.cell(row, PropID_Android_col).value
            PropID_Value = sheet.cell(row, PropID_Value_col).value
            ValueType = sheet.cell(row, ValueType_Android_col).value
            AreaType = sheet.cell(row, AreaType_Android_col).value
            ChangeMode = sheet.cell(row, ChangeMode_col).value
            Access = sheet.cell(row, Access_col).value
            v_Description = sheet.cell(row, Value_Description_col).value
            f_Description = sheet.cell(row, Function_Description_col).value
            if propID == PropID_Android:
                if PropID_AreaType_dict[PropID_Android]['PropID_Value'] == '':
                    PropID_AreaType_dict[PropID_Android]['PropID_Value'] = PropID_Value
                elif PropID_AreaType_dict[PropID_Android]['PropID_Value'] != PropID_Value:
                    print('[error] same PropID_Android has diff PropID_Value: ' + PropID_Android)
                if PropID_AreaType_dict[PropID_Android]['ValueType'] == '':
                    PropID_AreaType_dict[PropID_Android]['ValueType'] = ValueType
                elif PropID_AreaType_dict[PropID_Android]['ValueType'] != ValueType:
                    print('[error] same PropID_Android has diff ValueType: ' + PropID_Android)
                if PropID_AreaType_dict[PropID_Android]['AreaType'] == '':
                    PropID_AreaType_dict[PropID_Android]['AreaType'] = AreaType
                elif PropID_AreaType_dict[PropID_Android]['AreaType'] != AreaType:
                    print('[error] same PropID_Android has diff AreaType: ' + PropID_Android)
                if PropID_AreaType_dict[PropID_Android]['ChangeMode'] == '':
                    PropID_AreaType_dict[PropID_Android]['ChangeMode'] = ChangeMode
                elif PropID_AreaType_dict[PropID_Android]['ChangeMode'] != ChangeMode:
                    print('[error] same PropID_Android has diff ChangeMode: ' + PropID_Android)
                if PropID_AreaType_dict[PropID_Android]['Access'] == '':
                    PropID_AreaType_dict[PropID_Android]['Access'] = Access
                elif PropID_AreaType_dict[PropID_Android]['Access'] != Access:
                    PropID_AreaType_dict[PropID_Android]['Access'] = 'READ_WRITE'
                if (v_Description == '' or v_Description is None):
                    PropID_AreaType_dict[PropID_Android]['v_Description'] = ''
                else:
                    Description = v_Description.replace('\n', "\n\t\t")
                    Description = '\n\t\t'+Description.strip()
                    PropID_AreaType_dict[PropID_Android]['v_Description'] = Description
                if (f_Description == '' or f_Description is None):
                    PropID_AreaType_dict[PropID_Android]['f_Description'] = ''
                else:
                    PropID_AreaType_dict[PropID_Android]['f_Description'] = f_Description

    # 生成主体代码
    for PropID_Android in propId_list:
        PropID_Value = PropID_AreaType_dict[PropID_Android]['PropID_Value']
        ValueType = PropID_AreaType_dict[PropID_Android]['ValueType']
        AreaType = PropID_AreaType_dict[PropID_Android]['AreaType']
        ChangeMode = PropID_AreaType_dict[PropID_Android]['ChangeMode']
        Access = PropID_AreaType_dict[PropID_Android]['Access']
        v_Description = PropID_AreaType_dict[PropID_Android]['v_Description'].replace('\r\n', ' <br>\r\n').replace('\n', ' <br>\n')
        f_Description = PropID_AreaType_dict[PropID_Android]['f_Description']
        g_vehicle_property_Ids_list +='''
    /**
     * %s <p>
     * <br>
     * Function Describe: %s <p>
     * VehiclePropertyChangeMode:%s <p>
     * VehiclePropertyAccess:%s <p>
     * Signal Value:%s <p>
     */
    ''' % (PropID_Android, f_Description, ChangeMode, Access, v_Description)
        property_id = int(PropID_Value, 16)
        property_id += int('0x20000000', 16) #VehiclePropertyGroup SYSTEM=0x10000000, VENDOR=0x20000000
        property_id += int('0x01000000', 16) #VehicleArea GLOBAL=0x01000000, WINDOW=0x03000000, MIRROR=0x04000000,SEAT=0x05000000, DOOR        = 0x06000000, WHEEL       = 0x07000000

        if ValueType == 'STRING':
            property_id += int('0x00100000', 16)
        elif ValueType == 'BOOL': #BOOLEAN
            property_id += int('0x00200000', 16)
        elif ValueType == 'INT32':
            property_id += int('0x00400000', 16)
        elif ValueType == 'INT32_VEC':
            property_id += int('0x00410000', 16)
        elif ValueType == 'INT64':
            property_id += int('0x00500000', 16)
        elif ValueType == 'INT64_VEC':
            property_id += int('0x00510000', 16)
        elif ValueType == 'FLOAT':
            property_id += int('0x00600000', 16)
        elif ValueType == 'FLOAT_VEC':
            property_id += int('0x00610000', 16)
        elif ValueType == 'BYTES':
            property_id += int('0x00700000', 16)
        else:
            print(f'[error] {PropID_Android} ValueType not find.')

        g_vehicle_property_Ids_list += '\tpublic static final int %s = %s;\n' % (PropID_Android, property_id)

##########
## 生成 VehicleHal_setVehiclePropValue.cpp 文件 数据
## 查询Android侧INT32_VEC信号
##
## 全局变量 g_android_write_int32vec_list
##
##
##########
def find_android_int32_vec(sheet):
    global g_android_write_int32vec_list
    PropID_Android_col = g_title_col_dict["PropID_Android"]
    ValueType_Android_col = g_title_col_dict["ValueType_Android"]
    Access_col = g_title_col_dict['Access']
    for row in range(MAP_COTENT_ROW, sheet.max_row + 1): #(3-last)行
        PropID_Android = sheet.cell(row, PropID_Android_col).value
        ValueType = sheet.cell(row, ValueType_Android_col).value
        Access = sheet.cell(row, Access_col).value
        if PropID_Android != '' and ValueType == 'INT32_VEC' and Access == 'WRITE':
            print('Android Write INT32_VEC Signal: %s' % PropID_Android)
            g_android_write_int32vec_list += 'propValue.prop == toInt(VendorVehicleProperty::%s)\n' % PropID_Android


##########
## 生成 Types_Config.cpp 文件 数据
## sheet: config 生成enum
##
##
##########
def generate_config_types_code(sheet, project):
    # 生成types.hal中areaid
    config_areaid_col = 0
    config_idvalue_col = 0
    config_areaType_col = 0
    config_areaTypeValue_col = 0

    for index in range(1, sheet.max_column + 1):#(1-last)列
        if sheet.cell(1, index).value == 'AreaID':
            config_areaid_col = index
        if sheet.cell(1, index).value == 'IDValue':
            config_idvalue_col = index

        if sheet.cell(1, index).value == 'AreaType':
            config_areaType_col = index
        if sheet.cell(1, index).value == 'AreaTypeValue':
            config_areaTypeValue_col = index

    VendorVehicleArea = ''
    VendorVehicleAreaBody = ''
    VendorVehicleAreaSeat = ''
    VendorVehicleAreaWindow = ''
    VendorVehicleAreaDoor = ''
    VendorVehicleAreaMirror = ''
    VendorVehicleAreaGlobal = ''

    for index in range(2, sheet.max_row + 1): #(2-last)行
        AreaType = sheet.cell(index, config_areaType_col).value
        AreaTypeValue = sheet.cell(index, config_areaTypeValue_col).value
        if AreaType is None or AreaType == '':
            break
        VendorVehicleArea += '\t%-30s = %s,\n' % (AreaType, AreaTypeValue)

    for index in range(2, sheet.max_row + 1): #(2-last)行
        AreaID = sheet.cell(index, config_areaid_col).value
        IDValue = sheet.cell(index, config_idvalue_col).value
        if(AreaID is None or AreaID == ''):
            break
        if 'GLOBAL' in AreaID:
            VendorVehicleAreaGlobal += '\t%-30s = %s,\n' % (AreaID, IDValue)
        elif 'SEAT' in AreaID:
            VendorVehicleAreaSeat += '\t%-30s = %s,\n' % (AreaID, IDValue)
        elif 'WINDOW' in AreaID:
            VendorVehicleAreaWindow += '\t%-30s = %s,\n' % (AreaID, IDValue)
        elif 'DOOR' in AreaID:
            VendorVehicleAreaDoor += '\t%-30s = %s,\n' % (AreaID, IDValue)
        elif 'MIRROR' in AreaID:
            VendorVehicleAreaMirror += '\t%-30s = %s,\n' % (AreaID, IDValue)
        elif 'BODY' in AreaID:
            VendorVehicleAreaBody += '\t%-30s = %s,\n' % (AreaID, IDValue)

    tmplate_dictionary = {'VendorVehicleArea': VendorVehicleArea.expandtabs(tabsize=4),
                          'VendorVehicleAreaBody': VendorVehicleAreaBody.expandtabs(tabsize=4),
                          'VendorVehicleAreaSeat': VendorVehicleAreaSeat.expandtabs(tabsize=4),
                          'VendorVehicleAreaMirror': VendorVehicleAreaMirror.expandtabs(tabsize=4),
                          'VendorVehicleAreaWindow': VendorVehicleAreaWindow.expandtabs(tabsize=4),
                          'VendorVehicleAreaDoor': VendorVehicleAreaDoor.expandtabs(tabsize=4),
                          'VendorVehicleAreaGlobal': VendorVehicleAreaGlobal.expandtabs(tabsize=4)
                          }
    return tmplate_dictionary

##########
## 获取所有入参标题所在列，生成列名及索引字典
##
##  生成全局字典 g_list_num_dict
##########
def read_title_col(sheet):
    global g_title_col_dict
    title_name = g_title_name_list.split(',')
    title_svw_name = g_title_name_svw_list.split(',')


    for n, title in enumerate(title_svw_name):
        for index in range(1, sheet.max_column + 1): # (1-last)列
            if sheet.cell(MAP_TITLE_ROW, index).value.startswith(title):
                tit_element = title_name[n]
                g_title_col_dict[tit_element] = index
                break
                
#    for title in title_name:
#        for index in range(1, sheet.max_column + 1): # (1-last)列
#            if sheet.cell(MAP_TITLE_ROW, index).value.startswith(title):
#                g_title_col_dict[title] = index
#                break

    for title in title_name:
        if title not in g_title_col_dict.keys():
            print('[error] not find title = %s in Map Sheet' % title)
    if(isDebug):
        print(g_title_col_dict)

##########
## 移除列表中重复元素
##########
def remove_same_prop_elements(L):
    T = []
    for i in L:
        if not i in T:
            T.append(i)
    return T

##########
## 生成文件工具
## tmplate_dictionary 数据字典
## template_name 模板文件名
## result_name (选填) 目标文件名 如果不填同模板文件夹名
##########
def make_result_resource_file(tmplate_dictionary, template_name, result_name=None):
    
    if result_name is None:
        result_path = 'Result' + project + '/' + template_name
        template_path = 'template/' + template_name
    else:
        result_path = 'Result' + project + '/' + result_name
        template_path = 'template/' + template_name

    print(f'result_path {result_path}')
    try:
        os.remove(result_path)
    except:
        print(f'no {result_path} file')

    try:
        # 确保目录存在
        dir_path = os.path.dirname(result_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path)

        # 打开文件（带错误处理）
        with open(template_path, 'r', encoding='utf-8') as fd:
            template = fd.read() # 读取整个文件内容到字符串变量
        fd = open(result_path, 'w+', encoding='utf-8')
        fd.write(template % tmplate_dictionary)
        fd.close()
        print(f"The {result_path} file has been created.")
    except FileNotFoundError:
        print(f"错误：目录 '{dir_path}' 不存在，请检查路径。")
    except PermissionError:
        print(f"错误：没有写入权限，请检查 '{result_path}' 的权限设置。")
    except OSError as e:
        print(f"操作系统错误：{e}")
    except Exception as e:
        print(f"未知错误：{e}")

def make_result_json_file(file_name, title, content_json):
    json_root = {}
    json_root[title] = content_json
    result_path = 'Result' + project + '/'  + file_name

    try:
        # 确保目录存在
        dir_path = os.path.dirname(result_path)
        if dir_path and not os.path.exists(dir_path):
            os.makedirs(dir_path)

        os.remove(result_path)
    except:
        print(f"no {result_path} file")
    fd = open(result_path, 'w+', encoding='utf-8')
    fd.write(json.dumps(json_root, indent=4))
    fd.close()
    print(f"The {file_name} file has been created.")

if __name__ == '__main__':
    excel_path = './' + sys.argv[1] # excel path
    project = sys.argv[2] # Project_Name
    sheet_name = sys.argv[3] # sheet name

    if (isDebug != True):
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    workbook = load_workbook(excel_path) # 加载已有的 Excel 文件
    sheet_map = workbook[sheet_name]# 通过工作表名称获取map矩阵工作表
    sheet_config = workbook['Config']

    if 1:
        #vehif_version = (sys.argv[1]).split('\\')[-1].split('.xls')[0]
        vehif_version = (sys.argv[1]).split('/')[-1].split('.xlsx')[0]
        # 定义正则表达式
        pattern = r'V\d+.\d+.\d+'

        # 使用re.search()查找匹配
        match = re.search(pattern, vehif_version)

        # 检查是否找到匹配并输出结果
        if match:
            vehif_version = match.group()  # 提取匹配的字符串
        else:
            vehif_version = 'V1.0.0'
#        print('自动生成的版本号：' + vehif_version)
    else:
        vehif_version = 'P12L_VehIF_V1.1.14'
#        print('手动输入的版本号：' + vehif_version)

    os.makedirs(os.path.dirname('Result' + project + '/'), exist_ok=True)
    # 获取标题所在列
    read_title_col(sheet_map)

    # 生成订阅代码、Signal_*.h、getPropValueID中使用
    enum_class = 'EViPropId_'+project

    # 生成VehicleIF缓存PropValue json文件  g_initPropJson_list
    if(g_need_creat_vehicle_prop_value_json):
        generate_initPropValue_json(sheet_map)
        make_result_json_file('vehicle_prop_value.json', 'propValue', g_initPropJson_list)

    # 生成VehicleIF PropConfig json文件
    if(g_need_creat_vehicle_prop_config_json):
        generate_initPropConfigJson_file(sheet_map)
        make_result_json_file('vehicle_prop_config.json', 'propConfig', g_initPropCfgJson_list)

    # 生成ViClient中Signal_*.h文件
    if(g_need_creat_Signal_H):
        generate_viclient_signal_code(sheet_map)
        tmplate_dictionary = {'g_propid_value_list': g_viclient_signal_list.expandtabs(tabsize=4), 'projectName': enum_class, 'macro': project}
        make_result_resource_file(tmplate_dictionary, 'Signal_project.h', 'Signal.h')

    # 生成ViClient中autolink.alc.signal.pb.proto文件
    if(g_need_creat_Signal_Proto):
        g_viclient_signal_list = ''
        generate_viclient_signal_code(sheet_map)
        g_viclient_signal_list = g_viclient_signal_list.replace(",", ";")
        tmplate_dictionary = {'g_propid_value_list': g_viclient_signal_list.expandtabs(tabsize=4), 'projectName': enum_class, 'macro': project}
        make_result_resource_file(tmplate_dictionary, 'autolink.alc.signal.pb.proto')


    # 生成VehicleIF缓存PropValue
    if(g_need_creat_InitVehiclePropValue_CPP):
        generate_initPropValue_code(sheet_map)
        tmplate_dictionary = {'vehicle_value_list': g_initPropValue_list.expandtabs(tabsize=4), 'version': vehif_version}
        make_result_resource_file(tmplate_dictionary, 'InitVehiclePropValue_New.cpp')

    # 生成VehicleIF缓存PropConfig
    if(g_need_creat_InitVehiclePropConfig_CPP):
        generate_initPropConfig_code(sheet_map)
        tmplate_dictionary = {'g_initPropConfig_list': g_initPropConfig_list.expandtabs(tabsize=4), 'version': vehif_version}
        make_result_resource_file(tmplate_dictionary, 'InitVehiclePropConfig_New.cpp')

    # 生成Android客户端订阅代码
    if(g_need_creat_SubscribeProps_CPP):
        generate_android_subscribe_code(sheet_map, enum_class)
        tmplate_dictionary = {'g_android_subscribe_list': g_android_subscribe_list.expandtabs(tabsize=4)}
        make_result_resource_file(tmplate_dictionary, 'SubscribeProps.cpp')

    # 生成QNX客户端订阅代码
    if(g_need_creat_QNX_Subscribe_New_CPP):
        generate_qnx_subscribe_code(sheet_map, enum_class)
        tmplate_dictionary = {'g_qnx_subscribe_list': g_qnx_subscribe_list.expandtabs(tabsize=4)}
        make_result_resource_file(tmplate_dictionary, 'QNX_Subscribe_New.cpp')

    # 生成VehicleHal中VehicleHalFunc::getPropValueID方法
    if(g_need_creat_GetPropValueID_CPP):
        generate_getPropValueID_code(sheet_map, enum_class)
        tmplate_dictionary = {'g_getPropValueID_list': g_getPropValueID_list.expandtabs(tabsize=4)}
        make_result_resource_file(tmplate_dictionary, 'GetPropValueID.cpp')

    # 生成VehicleHal PropConfig
    if(g_need_creat_InitPropConfigs_CPP):
        generate_android_propConfig_code(sheet_map)
        tmplate_dictionary = {'g_android_propConfig_list': g_android_propConfig_list.expandtabs(tabsize=4)}
        make_result_resource_file(tmplate_dictionary, 'InitPropConfigs_' + project + '.cpp', 'InitPropConfigs.cpp')

    # 生成Types.hal
    if(g_need_creat_Types_Hal_New):
        generate_types_hal_code(sheet_map)
        tmplate_dictionary = {'g_types_hal_list': g_types_hal_list.expandtabs(tabsize=4), 'projectName': project}
        make_result_resource_file(tmplate_dictionary, 'hw_types.hal')

        generate_config_types_code(sheet_config, project)
        tmplate_dictionary.update(generate_config_types_code(sheet_config, project))
        tmplate_dictionary['projectName'] = project
        if(project == 'N80'):
            tmplate_dictionary['version'] = '2.1'
        elif(project == 'N50'):
            tmplate_dictionary['version'] = '2.0'
        else:
            tmplate_dictionary['version'] = '2.0'
        make_result_resource_file(tmplate_dictionary, 'hal_types.hal')

    # 生成ALVehiclePropertyIds.java
    if(g_need_creat_ALVehiclePropertyIds_JAVA):
        generate_ALVehiclePropertyIds_code(sheet_map)
        tmplate_dictionary = {'g_vehicle_property_Ids_list': g_vehicle_property_Ids_list.expandtabs(tabsize=4), 'projectName': project}
        make_result_resource_file(tmplate_dictionary, 'ALVehiclePropertyIds.java', 'AL'+ project +'VehiclePropertyIds.java')

    # 查询Android侧INT32_VEC信号
    if(g_need_creat_VehicleHal_setVehiclePropValue_CPP):
        find_android_int32_vec(sheet_map)
        tmplate_dictionary = {'android_int32vec_prop': g_android_write_int32vec_list.expandtabs(tabsize=4)}
        make_result_resource_file(tmplate_dictionary, 'VehicleHal_setVehiclePropValue.cpp')
