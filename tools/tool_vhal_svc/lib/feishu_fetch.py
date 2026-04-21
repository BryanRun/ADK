"""飞书电子表格导出为 xlsx（与 tool_property 同源逻辑，供 vhal-svc fetch 使用）。"""

from __future__ import annotations

import os
import time

FEISHU_BASE = "https://open.feishu.cn/open-apis"

# Sheets API 降级重建 xlsx 时，按 metainfo 的 columnCount 拉全表列（上限对齐 Excel 最大列 XFD）。
# 旧实现 min(col, 50) 会导致第 51 列及以后丢失（如 AreaID_VHal 靠后即无法匹配表头）。
_MAX_SHEETS_API_COLS = 16384


def require_requests():
    try:
        import requests

        return requests
    except ImportError as e:
        raise RuntimeError("fetch 需要 requests 库（与平台 pyproject 一致）") from e


def get_tenant_access_token() -> str | None:
    app_id = os.environ.get("FEISHU_APP_ID", "").strip()
    app_secret = os.environ.get("FEISHU_APP_SECRET", "").strip()
    if not app_id or not app_secret:
        return None
    requests = require_requests()
    r = requests.post(
        f"{FEISHU_BASE}/auth/v3/tenant_access_token/internal",
        json={"app_id": app_id, "app_secret": app_secret},
        timeout=60,
    )
    data = r.json()
    if data.get("code") != 0:
        return None
    return data.get("tenant_access_token")


def parse_feishu_spreadsheet_url(value: str) -> tuple[str, str]:
    """返回 (token, url_type)。url_type 为 wiki / sheets / raw。"""
    value = (value or "").strip()
    if not value or "/" not in value:
        return value, "raw"
    path = value.split("?")[0].rstrip("/")
    if "/wiki/" in value:
        return path.rsplit("/", 1)[-1], "wiki"
    if "/sheets/" in value:
        return path.rsplit("/", 1)[-1], "sheets"
    return path.rsplit("/", 1)[-1], "raw"


def get_spreadsheet_title(
    requests, headers: dict, spreadsheet_token: str
) -> str | None:
    """飞书表格标题（与在线文档名称一致，用于本地 xlsx 文件名）。"""
    r = requests.get(
        f"{FEISHU_BASE}/sheets/v3/spreadsheets/{spreadsheet_token}",
        headers=headers,
        timeout=60,
    )
    data = r.json()
    if data.get("code") != 0:
        return None
    sp = data.get("data", {}).get("spreadsheet") or {}
    t = sp.get("title")
    return str(t).strip() if t else None


def sanitize_filename_from_title(title: str, *, fallback: str = "vhal_matrix") -> str:
    """将表格标题转为安全文件名（不含扩展名）。"""
    if not title or not str(title).strip():
        return fallback
    t = str(title).strip()
    for c in '\\/:*?"<>|\n\r\t':
        t = t.replace(c, "_")
    t = t.strip(". ")
    if not t:
        return fallback
    if len(t) > 180:
        t = t[:180]
    return t


def resolve_wiki_to_spreadsheet_token(requests, headers: dict, wiki_token: str) -> str | None:
    r = requests.get(
        f"{FEISHU_BASE}/wiki/v2/spaces/get_node",
        headers=headers,
        params={"token": wiki_token},
        timeout=60,
    )
    data = r.json()
    if data.get("code") != 0:
        return None
    node = data.get("data", {}).get("node", {})
    if node.get("obj_type") != "sheet":
        return None
    return node.get("obj_token") or None


def export_xlsx(requests, headers: dict, spreadsheet_token: str, timeout: int = 120) -> bytes | None:
    r = requests.post(
        f"{FEISHU_BASE}/drive/v1/export_tasks",
        headers=headers,
        json={"file_extension": "xlsx", "token": spreadsheet_token, "type": "sheet"},
        timeout=60,
    )
    data = r.json()
    if data.get("code") != 0:
        return None
    ticket = data.get("data", {}).get("ticket")
    if not ticket:
        return None
    elapsed = 0
    while elapsed < timeout:
        time.sleep(2)
        elapsed += 2
        r = requests.get(
            f"{FEISHU_BASE}/drive/v1/export_tasks/{ticket}",
            headers=headers,
            params={"token": spreadsheet_token},
            timeout=60,
        )
        result = r.json().get("data", {}).get("result", {})
        job_status = result.get("job_status", -1)
        if job_status == 0:
            r = requests.get(
                f"{FEISHU_BASE}/drive/v1/export_tasks/file/{ticket}",
                headers=headers,
                params={"token": spreadsheet_token},
                timeout=120,
            )
            if r.status_code == 200 and len(r.content) > 0:
                return r.content
            return None
        if job_status not in (1, 2):
            return None
    return None


def _col_letter(col_num: int) -> str:
    result = ""
    while col_num > 0:
        col_num, rem = divmod(col_num - 1, 26)
        result = chr(65 + rem) + result
    return result


def download_via_sheets_api(requests, headers: dict, spreadsheet_token: str) -> bytes | None:
    from io import BytesIO

    from openpyxl import Workbook

    r = requests.get(
        f"{FEISHU_BASE}/sheets/v2/spreadsheets/{spreadsheet_token}/metainfo",
        headers=headers,
        timeout=60,
    )
    data = r.json()
    if data.get("code") != 0:
        return None
    sheets = data["data"]["sheets"]
    wb = Workbook()
    wb.remove(wb.active)
    for idx, sheet_info in enumerate(sheets):
        title = sheet_info["title"]
        sheet_id = sheet_info["sheetId"]
        row_count = int(sheet_info.get("rowCount") or 0)
        col_count = int(sheet_info.get("columnCount") or 0)
        n_cols = min(max(col_count, 1), _MAX_SHEETS_API_COLS)
        n_rows = max(row_count, 1)
        end_col = _col_letter(n_cols)
        range_str = f"{sheet_id}!A1:{end_col}{n_rows}"
        r = requests.get(
            f"{FEISHU_BASE}/sheets/v2/spreadsheets/{spreadsheet_token}/values/{range_str}",
            headers=headers,
            params={"valueRenderOption": "ToString"},
            timeout=120,
        )
        resp = r.json()
        if resp.get("code") != 0:
            ws = wb.create_sheet(title=title)
            continue
        values = resp.get("data", {}).get("valueRange", {}).get("values", [])
        ws = wb.create_sheet(title=title)
        for row_idx, row in enumerate(values, 1):
            if row is None:
                continue
            for col_idx, val in enumerate(row, 1):
                if val is not None:
                    ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=str(val) if not isinstance(val, (int, float)) else val,
                    )
        if (idx + 1) % 5 == 0:
            print(f"    {idx + 1}/{len(sheets)} sheets...")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fetch_spreadsheet_to_bytes_named(
    token_or_url: str, *, access_token: str | None = None
) -> tuple[bytes | None, str, str]:
    """返回 (xlsx 内容, 建议本地文件名含 .xlsx, 错误说明)。文件名与飞书表格标题一致（经安全化）。"""
    requests = require_requests()
    token = access_token or get_tenant_access_token()
    if not token:
        return None, "", "未设置 FEISHU_APP_ID / FEISHU_APP_SECRET 或获取 token 失败"

    headers = {"Authorization": f"Bearer {token}"}
    stoken, url_type = parse_feishu_spreadsheet_url(token_or_url)
    spreadsheet_token = stoken
    if url_type == "wiki":
        resolved = resolve_wiki_to_spreadsheet_token(requests, headers, stoken)
        if not resolved:
            return None, "", "Wiki 节点解析失败或节点不是电子表格"
        spreadsheet_token = resolved
        print(f"  Wiki → spreadsheet token: {spreadsheet_token[:12]}...")

    title = get_spreadsheet_title(requests, headers, spreadsheet_token)
    base = sanitize_filename_from_title(title or "")
    filename = f"{base}.xlsx"
    if title:
        print(f"  飞书表格标题: {title}")
        print(f"  本地文件名: {filename}")

    content = export_xlsx(requests, headers, spreadsheet_token)
    if content is None:
        print("  Drive Export API 不可用或超时，尝试 Sheets API 重建...")
        content = download_via_sheets_api(requests, headers, spreadsheet_token)
        if content:
            print("  Sheets API 重建完成")
    if content is None:
        return None, filename, "导出 xlsx 失败"
    return content, filename, ""


def fetch_spreadsheet_to_bytes(
    token_or_url: str, *, access_token: str | None = None
) -> tuple[bytes | None, str]:
    """返回 (xlsx 内容, 错误说明)。不保证文件名与标题一致，请优先使用 fetch_spreadsheet_to_bytes_named。"""
    content, _, err = fetch_spreadsheet_to_bytes_named(
        token_or_url, access_token=access_token
    )
    return content, err
