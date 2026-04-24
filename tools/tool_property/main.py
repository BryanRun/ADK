#!/usr/bin/env python3
"""
property — 车辆属性（CarProperty）生成与导入

实现目录：tools/tool_property/（对外 CLI：adk property）

Usage:
    python main.py                        # 全部项目：scan + fetch + generate + deploy + snapshot
    python main.py BAIC                   # 仅 BAIC：scan + fetch + generate + deploy + snapshot
    python main.py fetch                  # 全部项目：仅 fetch
    python main.py -f                     # 同上
    python main.py generate BAIC          # 仅 BAIC：仅 generate
    python main.py -g BAIC               # 同上
    python main.py scan BAIC             # 仅 BAIC：仅 scan（对比飞书与本地差异）
    python main.py snapshot BAIC         # 仅 BAIC：仅创建版本快照
    python main.py -fgd BAIC             # 仅 BAIC：fetch + generate + deploy
    python main.py list                   # 列出所有已配置项目
    python main.py -l                     # 同上

操作（scan/fetch/generate/deploy/snapshot）可自由组合，不指定时默认执行完整流水线。
Excel 表格放入 input/<项目名>/ 目录，生成结果输出到 output/<项目名>/ 目录。
"""

import os
import sys
import re
import glob
import json
import time
import shutil
import argparse
import subprocess
from datetime import datetime

VERSION = '1.2.0'

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from lib.term_color import green, red, yellow, bold
from model.property_model import PropertyGenerator

try:
    from autodrivekit.tool_runtime_paths import ensure_tool_config_seeded, resolve_tool_paths

    _PATHS = resolve_tool_paths(__file__)
    ensure_tool_config_seeded(_PATHS, "config.json")
    CONFIG_FILE = str(_PATHS.config_file("config.json"))
    INPUT_DIR = str(_PATHS.workspace_sub("input"))
    OUTPUT_DIR = str(_PATHS.workspace_sub("output"))
    TEMPLATE_DIR = os.path.join(str(_PATHS.bundle_dir), "templates")
    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
except ImportError:
    _PATHS = None
    CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.json")
    INPUT_DIR = os.path.join(SCRIPT_DIR, "input")
    OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
    TEMPLATE_DIR = os.path.join(SCRIPT_DIR, "templates")


def load_config():
    if not os.path.isfile(CONFIG_FILE):
        print(f"错误: 配置文件不存在 - {CONFIG_FILE}")
        sys.exit(1)
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def run_git(repo_path, *args):
    """在指定仓库执行 git 命令，返回 (成功?, stdout, stderr)。"""
    cmd = ['git', '-C', repo_path] + list(args)
    result = subprocess.run(cmd, capture_output=True, text=True)
    return result.returncode == 0, result.stdout.strip(), result.stderr.strip()


def extract_date_from_filename(filepath):
    """从文件名中提取日期，用于版本比较。"""
    basename = os.path.splitext(os.path.basename(filepath))[0]

    m = re.search(r'(\d{4})[-_](\d{2})[-_](\d{2})', basename)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    m = re.search(r'(\d{4})(\d{2})(\d{2})', basename)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    for m in re.finditer(r'(\d{2})(\d{2})(\d{2})', basename):
        yy, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mm <= 12 and 1 <= dd <= 31:
            try:
                return datetime(2000 + yy, mm, dd)
            except ValueError:
                continue

    return datetime.fromtimestamp(os.path.getmtime(filepath))


def find_latest_excel(project_input_dir):
    """在项目 input 目录下找到日期最新的 .xlsx 文件。"""
    xlsx_files = glob.glob(os.path.join(project_input_dir, '*.xlsx'))
    if not xlsx_files:
        return None

    best_file = None
    best_date = None
    for f in xlsx_files:
        date = extract_date_from_filename(f)
        if best_date is None or date > best_date:
            best_date = date
            best_file = f

    return best_file


def ensure_project_dirs(projects):
    """为所有配置的项目创建 input 子目录（如不存在）。"""
    for name in projects:
        d = os.path.join(INPUT_DIR, name)
        os.makedirs(d, exist_ok=True)


# ---------------------------------------------------------------------------
#  飞书 API：fetch 子命令所需
# ---------------------------------------------------------------------------

FEISHU_BASE = 'https://open.feishu.cn/open-apis'
# Sheets API 重建 xlsx 时按 metainfo 全列读取（上限 Excel XFD）；勿再用 50 列以免靠后列丢失
FEISHU_SHEETS_REBUILD_MAX_COLS = 16384


def _require_requests():
    try:
        import requests
        return requests
    except ImportError:
        print("错误: fetch 功能需要 requests 库")
        print("  安装: pip install requests")
        sys.exit(1)


def get_feishu_token():
    """通过环境变量获取飞书 tenant_access_token。"""
    app_id = os.environ.get('FEISHU_APP_ID', '')
    app_secret = os.environ.get('FEISHU_APP_SECRET', '')
    if not app_id or not app_secret:
        print("错误: 飞书 API 凭证未设置")
        print("  请设置环境变量:")
        print("    export FEISHU_APP_ID='your_app_id'")
        print("    export FEISHU_APP_SECRET='your_app_secret'")
        return None

    requests = _require_requests()
    r = requests.post(f'{FEISHU_BASE}/auth/v3/tenant_access_token/internal', json={
        'app_id': app_id, 'app_secret': app_secret
    })
    data = r.json()
    if data.get('code') != 0:
        print(f"错误: 获取飞书 token 失败 - {data.get('msg', '未知错误')}")
        return None
    return data['tenant_access_token']


def _feishu_create_version(requests, headers, spreadsheet_token, version_name):
    """在飞书表格上创建命名版本快照。"""
    r = requests.post(
        f'{FEISHU_BASE}/drive/v1/files/{spreadsheet_token}/versions',
        headers=headers,
        json={'name': version_name, 'obj_type': 'sheet'}
    )
    data = r.json()
    if data.get('code') == 0:
        print(f"  版本快照: {version_name}")
        return True
    print(f"  版本快照: 创建失败 ({data.get('msg', '未知错误')})")
    return False


def _feishu_export_xlsx(requests, headers, spreadsheet_token, timeout=60):
    """异步导出飞书表格为 xlsx，返回二进制内容或 None。"""
    r = requests.post(
        f'{FEISHU_BASE}/drive/v1/export_tasks',
        headers=headers,
        json={'file_extension': 'xlsx', 'token': spreadsheet_token, 'type': 'sheet'}
    )
    data = r.json()
    if data.get('code') != 0:
        return None

    ticket = data['data']['ticket']

    elapsed = 0
    while elapsed < timeout:
        time.sleep(2)
        elapsed += 2
        r = requests.get(
            f'{FEISHU_BASE}/drive/v1/export_tasks/{ticket}',
            headers=headers,
            params={'token': spreadsheet_token}
        )
        result = r.json().get('data', {}).get('result', {})
        job_status = result.get('job_status', -1)
        if job_status == 0:
            file_token = result.get('file_token', '')
            if not file_token:
                return None
            r = requests.get(
                f'{FEISHU_BASE}/drive/v1/export_tasks/file/{ticket}',
                headers=headers,
                params={'token': spreadsheet_token}
            )
            if r.status_code == 200 and len(r.content) > 0:
                return r.content
            return None
        elif job_status in (1, 2):
            continue
        else:
            return None

    return None


def _col_letter(col_num):
    """列号(1-based)转 Excel 字母，如 1→A, 27→AA。"""
    result = ''
    while col_num > 0:
        col_num, rem = divmod(col_num - 1, 26)
        result = chr(65 + rem) + result
    return result


def _feishu_download_via_sheets_api(requests, headers, spreadsheet_token):
    """通过 Sheets API 逐 sheet 读取数据并用 openpyxl 重建 xlsx（导出 API 的降级方案）。"""
    from io import BytesIO
    from openpyxl import Workbook

    r = requests.get(
        f'{FEISHU_BASE}/sheets/v2/spreadsheets/{spreadsheet_token}/metainfo',
        headers=headers
    )
    data = r.json()
    if data.get('code') != 0:
        print(f"  ✗ 获取表格元数据失败: {data.get('msg', '未知错误')}")
        return None

    sheets = data['data']['sheets']
    wb = Workbook()
    wb.remove(wb.active)

    for idx, sheet_info in enumerate(sheets):
        title = sheet_info['title']
        sheet_id = sheet_info['sheetId']
        row_count = int(sheet_info.get('rowCount') or 0)
        col_count = int(sheet_info.get('columnCount') or 0)
        n_cols = min(max(col_count, 1), FEISHU_SHEETS_REBUILD_MAX_COLS)
        n_rows = max(row_count, 1)
        end_col = _col_letter(n_cols)
        range_str = f'{sheet_id}!A1:{end_col}{n_rows}'

        r = requests.get(
            f'{FEISHU_BASE}/sheets/v2/spreadsheets/{spreadsheet_token}/values/{range_str}',
            headers=headers,
            params={'valueRenderOption': 'ToString'}
        )
        resp = r.json()
        if resp.get('code') != 0:
            print(f"  ✗ 读取 sheet '{title}' 失败: {resp.get('msg', '未知错误')}")
            ws = wb.create_sheet(title=title)
            continue

        values = resp.get('data', {}).get('valueRange', {}).get('values', [])
        ws = wb.create_sheet(title=title)
        for row_idx, row in enumerate(values, 1):
            if row is None:
                continue
            for col_idx, val in enumerate(row, 1):
                if val is not None:
                    ws.cell(row=row_idx, column=col_idx, value=str(val) if not isinstance(val, (int, float)) else val)

        if (idx + 1) % 5 == 0:
            print(f"    {idx + 1}/{len(sheets)} sheets...")

    buf = BytesIO()
    wb.save(buf)
    print(f"  Sheets API 重建完成 ({len(sheets)} sheets)")
    return buf.getvalue()


def _parse_feishu_url(value):
    """解析飞书 URL，返回 (token, url_type)。url_type 为 'wiki'/'sheets'/'raw'。"""
    value = value.strip()
    if '/' not in value:
        return value, 'raw'

    path = value.split('?')[0].rstrip('/')
    if '/wiki/' in value:
        return path.rsplit('/', 1)[-1], 'wiki'
    if '/sheets/' in value:
        return path.rsplit('/', 1)[-1], 'sheets'
    return path.rsplit('/', 1)[-1], 'raw'


def _resolve_wiki_to_spreadsheet(requests, headers, wiki_token):
    """将 wiki node token 解析为实际的 spreadsheet obj_token。"""
    r = requests.get(
        f'{FEISHU_BASE}/wiki/v2/spaces/get_node',
        headers=headers,
        params={'token': wiki_token}
    )
    data = r.json()
    if data.get('code') != 0:
        print(f"  ✗ Wiki 节点解析失败: {data.get('msg', '未知错误')}")
        return None

    node = data.get('data', {}).get('node', {})
    obj_type = node.get('obj_type', '')
    obj_token = node.get('obj_token', '')

    if obj_type != 'sheet':
        print(f"  ✗ Wiki 节点不是电子表格 (类型: {obj_type})")
        return None

    if not obj_token:
        print(f"  ✗ 无法获取 spreadsheet token")
        return None

    print(f"  Wiki 解析: {wiki_token[:12]}... → {obj_token[:12]}...")
    return obj_token


def _resolve_spreadsheet_token(project_name, project_config, requests, headers):
    """解析项目配置中的 spreadsheet_token 并返回 (spreadsheet_token, sheet_title) 或 (None, None)。"""
    raw_token = project_config.get('spreadsheet_token', '')
    if not raw_token:
        print(f"  跳过: 未配置 spreadsheet_token")
        print(f"  提示: 在 config.json 中为 {project_name} 添加:")
        print(f'    "spreadsheet_token": "飞书表格 URL 或 token"')
        print(f'    支持格式: https://xxx.feishu.cn/sheets/<token>')
        print(f'              https://xxx.feishu.cn/wiki/<token>')
        return None, None

    token, url_type = _parse_feishu_url(raw_token)
    if url_type == 'wiki':
        spreadsheet_token = _resolve_wiki_to_spreadsheet(requests, headers, token)
        if spreadsheet_token is None:
            return None, None
    else:
        spreadsheet_token = token

    r = requests.get(f'{FEISHU_BASE}/sheets/v3/spreadsheets/{spreadsheet_token}', headers=headers)
    resp = r.json()
    if resp.get('code') == 0:
        sheet_title = resp['data']['spreadsheet']['title']
    else:
        sheet_title = project_config.get('excel_pattern', project_name.lower())
        print(f"  警告: 无法获取表格标题，使用 excel_pattern 作为文件名")

    return spreadsheet_token, sheet_title


def snapshot_project(project_name, project_config, feishu_token):
    """为飞书在线表格创建命名版本快照（独立命令）。"""
    requests = _require_requests()
    headers = {'Authorization': f'Bearer {feishu_token}'}

    spreadsheet_token, sheet_title = _resolve_spreadsheet_token(
        project_name, project_config, requests, headers)
    if spreadsheet_token is None:
        return False

    today_str = datetime.now().strftime('%Y%m%d')
    return _feishu_create_version(requests, headers, spreadsheet_token,
                                  f'{sheet_title}_{today_str}')


# ---------------------------------------------------------------------------
#  scan：飞书在线表格与本地 Excel 逐单元格对比
# ---------------------------------------------------------------------------

def _feishu_list_versions(requests, headers, spreadsheet_token):
    """获取飞书表格的版本快照列表，返回列表（按创建时间倒序）或空列表。"""
    items = []
    page_token = ''
    while True:
        params = {'obj_type': 'sheet', 'page_size': '50'}
        if page_token:
            params['page_token'] = page_token
        r = requests.get(
            f'{FEISHU_BASE}/drive/v1/files/{spreadsheet_token}/versions',
            headers=headers,
            params=params
        )
        data = r.json()
        if data.get('code') != 0:
            break
        page_data = data.get('data', {})
        for v in page_data.get('items', []):
            items.append(v)
        if not page_data.get('has_more'):
            break
        page_token = page_data.get('page_token', '')
        if not page_token:
            break
    return items


def _extract_date_str_from_name(name):
    """从版本快照名称或文件名中提取 YYYYMMDD 日期字符串。

    复用与 extract_date_from_filename 相同的匹配策略和有效性校验。
    """
    # 优先匹配 YYYY-MM-DD 或 YYYY_MM_DD
    for m in re.finditer(r'(\d{4})[-_](\d{2})[-_](\d{2})', name):
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= y <= 2099 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f'{y:04d}{mo:02d}{d:02d}'

    # 再匹配连续 YYYYMMDD
    for m in re.finditer(r'(\d{4})(\d{2})(\d{2})', name):
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= y <= 2099 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f'{y:04d}{mo:02d}{d:02d}'

    return None


def _feishu_read_sheets_to_dict(requests, headers, spreadsheet_token):
    """通过 Sheets API 读取飞书在线表格全部数据到内存，返回 {sheet_title: [[cell, ...], ...]}。"""
    r = requests.get(
        f'{FEISHU_BASE}/sheets/v2/spreadsheets/{spreadsheet_token}/metainfo',
        headers=headers
    )
    data = r.json()
    if data.get('code') != 0:
        print(f"  ✗ 获取表格元数据失败: {data.get('msg', '未知错误')}")
        return None

    result = {}
    sheets = data['data']['sheets']
    for idx, sheet_info in enumerate(sheets):
        title = sheet_info['title']
        sheet_id = sheet_info['sheetId']
        row_count = int(sheet_info.get('rowCount') or 0)
        col_count = int(sheet_info.get('columnCount') or 0)
        n_cols = min(max(col_count, 1), FEISHU_SHEETS_REBUILD_MAX_COLS)
        n_rows = max(row_count, 1)
        end_col = _col_letter(n_cols)
        range_str = f'{sheet_id}!A1:{end_col}{n_rows}'

        r = requests.get(
            f'{FEISHU_BASE}/sheets/v2/spreadsheets/{spreadsheet_token}/values/{range_str}',
            headers=headers,
            params={'valueRenderOption': 'ToString'}
        )
        resp = r.json()
        if resp.get('code') != 0:
            print(f"  ✗ 读取 sheet '{title}' 失败: {resp.get('msg', '未知错误')}")
            result[title] = []
            continue

        raw_values = resp.get('data', {}).get('valueRange', {}).get('values', [])
        rows = []
        for row in raw_values:
            if row is None:
                rows.append([])
            else:
                rows.append([_normalize_cell(v) for v in row])
        # 裁剪尾部空行（与本地 Excel 读取保持一致）
        while rows and all(c == '' for c in rows[-1]):
            rows.pop()
        result[title] = rows

        if (idx + 1) % 5 == 0:
            print(f"    {idx + 1}/{len(sheets)} sheets...")

    print(f"  在线读取完成 ({len(sheets)} sheets)")
    return result


def _read_local_excel_to_dict(excel_path):
    """用 openpyxl 读取本地 Excel 文件，返回与飞书读取相同格式的 {sheet_title: [[cell, ...], ...]}。"""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)
    result = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            cells = [_normalize_cell(cell.value) for cell in row]
            rows.append(cells)
        # 裁剪尾部空行
        while rows and all(c == '' for c in rows[-1]):
            rows.pop()
        result[ws.title] = rows
    wb.close()
    return result


def _normalize_cell(val):
    """归一化单元格值为字符串，便于对比。"""
    if val is None:
        return ''
    s = str(val).strip()
    try:
        n = float(s)
        if n == int(n):
            return str(int(n))
        return s
    except (ValueError, TypeError):
        return s


def _diff_sheets(old_data, new_data):
    """逐 sheet 逐单元格对比，返回结构化差异。

    返回: list of dict, 每个 dict 描述一个 sheet 的差异:
        {
            'sheet': str,
            'status': 'added' | 'removed' | 'changed' | 'unchanged',
            'added_rows': [(row_idx, [cell, ...]), ...],
            'removed_rows': [(row_idx, [cell, ...]), ...],
            'modified_cells': [(row_idx, col_idx, old_val, new_val), ...],
        }
    """
    all_sheets = list(dict.fromkeys(list(old_data.keys()) + list(new_data.keys())))
    diffs = []

    for sheet in all_sheets:
        if sheet not in old_data:
            new_rows = new_data[sheet]
            diffs.append({
                'sheet': sheet,
                'status': 'added',
                'added_rows': [(i, row) for i, row in enumerate(new_rows)],
                'removed_rows': [],
                'modified_cells': [],
            })
            continue

        if sheet not in new_data:
            old_rows = old_data[sheet]
            diffs.append({
                'sheet': sheet,
                'status': 'removed',
                'added_rows': [],
                'removed_rows': [(i, row) for i, row in enumerate(old_rows)],
                'modified_cells': [],
            })
            continue

        old_rows = old_data[sheet]
        new_rows = new_data[sheet]
        max_rows = max(len(old_rows), len(new_rows))
        max_cols = 0
        for r in old_rows + new_rows:
            if len(r) > max_cols:
                max_cols = len(r)

        added = []
        removed = []
        modified = []

        for i in range(max_rows):
            if i >= len(old_rows):
                added.append((i, new_rows[i]))
                continue
            if i >= len(new_rows):
                removed.append((i, old_rows[i]))
                continue
            old_r = old_rows[i]
            new_r = new_rows[i]
            row_max_cols = max(len(old_r), len(new_r))
            for j in range(row_max_cols):
                ov = old_r[j] if j < len(old_r) else ''
                nv = new_r[j] if j < len(new_r) else ''
                if ov != nv:
                    modified.append((i, j, ov, nv))

        status = 'unchanged'
        if added or removed or modified:
            status = 'changed'

        diffs.append({
            'sheet': sheet,
            'status': status,
            'added_rows': added,
            'removed_rows': removed,
            'modified_cells': modified,
        })

    return diffs


def _print_scan_report(project_name, diffs):
    """输出带颜色高亮的差异报告，返回是否有差异。"""
    has_diff = False
    total_added = 0
    total_removed = 0
    total_modified = 0
    changed_sheets = 0

    print(f"\n  [{project_name}] scan 结果:\n")

    for d in diffs:
        sheet = d['sheet']
        status = d['status']

        if status == 'added':
            has_diff = True
            changed_sheets += 1
            n = len(d['added_rows'])
            total_added += n
            print(green(f"    Sheet: {sheet}  [新增 sheet — {n} 行]"))
            for row_idx, row in d['added_rows'][:10]:
                preview = ' | '.join(row[:6])
                if len(row) > 6:
                    preview += ' | ...'
                print(green(f"      行 {row_idx + 1}: {preview}"))
            if n > 10:
                print(green(f"      ... 共 {n} 行"))
            print()
            continue

        if status == 'removed':
            has_diff = True
            changed_sheets += 1
            n = len(d['removed_rows'])
            total_removed += n
            print(red(f"    Sheet: {sheet}  [已删除 sheet — {n} 行]"))
            for row_idx, row in d['removed_rows'][:10]:
                preview = ' | '.join(row[:6])
                if len(row) > 6:
                    preview += ' | ...'
                print(red(f"      行 {row_idx + 1}: {preview}"))
            if n > 10:
                print(red(f"      ... 共 {n} 行"))
            print()
            continue

        if status == 'unchanged':
            print(f"    Sheet: {sheet}  — 无变更")
            continue

        # changed
        has_diff = True
        changed_sheets += 1
        print(f"    Sheet: {sheet}")

        added_rows = d['added_rows']
        removed_rows = d['removed_rows']
        modified_cells = d['modified_cells']

        if added_rows:
            total_added += len(added_rows)
            print(green(f"      新增行: {len(added_rows)} 行"))
            for row_idx, row in added_rows[:10]:
                preview = ' | '.join(row[:6])
                if len(row) > 6:
                    preview += ' | ...'
                print(green(f"        行 {row_idx + 1}: {preview}"))
            if len(added_rows) > 10:
                print(green(f"        ... 共 {len(added_rows)} 行"))

        if removed_rows:
            total_removed += len(removed_rows)
            print(red(f"      删除行: {len(removed_rows)} 行"))
            for row_idx, row in removed_rows[:10]:
                preview = ' | '.join(row[:6])
                if len(row) > 6:
                    preview += ' | ...'
                print(red(f"        行 {row_idx + 1}: {preview}"))
            if len(removed_rows) > 10:
                print(red(f"        ... 共 {len(removed_rows)} 行"))

        if modified_cells:
            total_modified += len(modified_cells)
            print(yellow(f"      修改: {len(modified_cells)} 处"))
            for row_idx, col_idx, ov, nv in modified_cells[:20]:
                col_letter = _col_letter(col_idx + 1)
                print(yellow(f"        行 {row_idx + 1}, 列 {col_letter}: {ov} → {nv}"))
            if len(modified_cells) > 20:
                print(yellow(f"        ... 共 {len(modified_cells)} 处"))

        print()

    # 汇总
    if has_diff:
        parts = []
        if total_added:
            parts.append(green(f"新增 {total_added} 行"))
        if total_removed:
            parts.append(red(f"删除 {total_removed} 行"))
        if total_modified:
            parts.append(yellow(f"修改 {total_modified} 处"))
        print(f"    汇总: {changed_sheets} 个 sheet 有变更，共 {' / '.join(parts)}")
    else:
        print(f"    与本地版本一致，无需更新。")

    return has_diff


def scan_project(project_name, project_config, feishu_token):
    """扫描飞书在线表格与本地 Excel 的差异。

    返回:
        'has_diff'   — 有差异
        'no_diff'    — 无差异
        'error'      — 出错
    """
    requests = _require_requests()
    headers = {'Authorization': f'Bearer {feishu_token}'}

    spreadsheet_token, sheet_title = _resolve_spreadsheet_token(
        project_name, project_config, requests, headers)
    if spreadsheet_token is None:
        return 'error'

    # --- 版本一致性校验 ---
    project_input = os.path.join(INPUT_DIR, project_name)
    local_excel = find_latest_excel(project_input) if os.path.isdir(project_input) else None

    versions = _feishu_list_versions(requests, headers, spreadsheet_token)
    latest_snapshot_date = None
    if versions:
        for v in versions:
            vname = v.get('name', '')
            d = _extract_date_str_from_name(vname)
            if d:
                latest_snapshot_date = d
                break

    local_date = None
    if local_excel:
        dt = extract_date_from_filename(local_excel)
        local_date = dt.strftime('%Y%m%d')

    version_mismatch = False
    if local_excel is None:
        version_mismatch = True
        print(bold(yellow(
            f"  ⚠ 本地 input/{project_name}/ 下无 Excel 文件"
        )))
        if latest_snapshot_date:
            print(bold(yellow(
                f"    飞书最新版本快照日期: {latest_snapshot_date}"
            )))
        print(bold(yellow(
            "    本地缺少对比基准，将以飞书当前内容作为全新增展示。"
        )))
        try:
            answer = input("    是否继续 scan？(y/n): ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            answer = 'n'
        if answer not in ('y', 'yes'):
            return 'error'
    elif latest_snapshot_date and local_date and latest_snapshot_date != local_date:
        version_mismatch = True
        print(bold(yellow(
            f"  ⚠ 版本日期不一致！"
        )))
        print(bold(yellow(
            f"    本地最新 Excel 日期: {local_date}  ({os.path.basename(local_excel)})"
        )))
        print(bold(yellow(
            f"    飞书最新快照日期:    {latest_snapshot_date}"
        )))
        print(bold(yellow(
            "    本地 Excel 与飞书最新快照不对应，对比结果可能不准确。"
        )))
        try:
            answer = input("    是否仍以本地 Excel 为基准继续 scan？(y/n): ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            answer = 'n'
        if answer not in ('y', 'yes'):
            return 'error'

    # --- 读取飞书在线数据 ---
    print(f"  正在从飞书读取在线表格数据...")
    new_data = _feishu_read_sheets_to_dict(requests, headers, spreadsheet_token)
    if new_data is None:
        return 'error'

    # --- 读取本地基准 ---
    if local_excel:
        print(f"  本地基准: {os.path.basename(local_excel)}")
        old_data = _read_local_excel_to_dict(local_excel)
    else:
        old_data = {}

    # --- 对比 ---
    diffs = _diff_sheets(old_data, new_data)
    has_diff = _print_scan_report(project_name, diffs)

    return 'has_diff' if has_diff else 'no_diff'


def fetch_project(project_name, project_config, feishu_token):
    """从飞书下载项目对应的电子表格到 input/ 目录（只读操作，不修改在线表格内容）。"""
    requests = _require_requests()
    headers = {'Authorization': f'Bearer {feishu_token}'}

    spreadsheet_token, sheet_title = _resolve_spreadsheet_token(
        project_name, project_config, requests, headers)
    if spreadsheet_token is None:
        return False

    today_str = datetime.now().strftime('%Y%m%d')
    filename = f'{sheet_title}_{today_str}.xlsx'
    save_path = os.path.join(INPUT_DIR, project_name, filename)

    # 导出并下载 xlsx（优先使用 Drive Export API，失败则降级为 Sheets API 重建）
    print(f"  导出中...")
    content = _feishu_export_xlsx(requests, headers, spreadsheet_token)
    if content is None:
        print(f"  Drive Export API 不可用，改用 Sheets API 读取...")
        content = _feishu_download_via_sheets_api(requests, headers, spreadsheet_token)
    if content is None:
        print(f"  ✗ 下载失败")
        return False

    # 3. 保存到 input/
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    exists_note = " (已覆盖同日版本)" if os.path.exists(save_path) else ""
    with open(save_path, 'wb') as f:
        f.write(content)

    size_kb = len(content) / 1024
    print(f"  已保存: input/{project_name}/{filename}  ({size_kb:.0f} KB){exists_note}")
    return True


def flatten_output(project_output):
    """通用逻辑：将生成引擎产出的单一子目录内容提升到项目根目录。"""
    subdirs = [d for d in os.listdir(project_output)
               if os.path.isdir(os.path.join(project_output, d))]
    for subdir in subdirs:
        src_dir = os.path.join(project_output, subdir)
        for fname in os.listdir(src_dir):
            shutil.move(os.path.join(src_dir, fname),
                        os.path.join(project_output, fname))
        os.rmdir(src_dir)


def expand_variants(project_output, variant_names):
    """特殊逻辑：按 variant_names 配置（输出名→内部源）生成变体目录。"""
    source_map = {}
    for output_name, source_name in variant_names.items():
        source_map.setdefault(source_name, []).append(output_name)

    for source, targets in source_map.items():
        src_path = os.path.join(project_output, source)
        if not os.path.isdir(src_path):
            continue

        first = targets[0]
        first_path = os.path.join(project_output, first)
        if first != source:
            if os.path.exists(first_path):
                shutil.rmtree(first_path)
            os.rename(src_path, first_path)

        for extra in targets[1:]:
            extra_path = os.path.join(project_output, extra)
            if os.path.exists(extra_path):
                shutil.rmtree(extra_path)
            shutil.copytree(first_path, extra_path)


def process_project(project_name, project_config):
    """处理单个项目：找最新表格 → 生成代码 → 后处理输出。"""
    project_input  = os.path.join(INPUT_DIR,  project_name)
    project_output = os.path.join(OUTPUT_DIR, project_name)

    if not os.path.isdir(project_input):
        print(f"  跳过: input/{project_name}/ 目录不存在")
        return False

    excel_path = find_latest_excel(project_input)
    if not excel_path:
        print(f"  跳过: input/{project_name}/ 下无 .xlsx 文件")
        return False

    date = extract_date_from_filename(excel_path)
    print(f"  表格: {os.path.basename(excel_path)}  (版本日期: {date.strftime('%Y-%m-%d')})")

    generator = PropertyGenerator(template_path=TEMPLATE_DIR, output_dir=project_output)
    generator.prop_process(excel_path)

    variant_names = project_config.get('variant_names')
    if variant_names:
        expand_variants(project_output, variant_names)
    else:
        flatten_output(project_output)

    # 输出摘要
    entries = sorted(os.listdir(project_output)) if os.path.isdir(project_output) else []
    subdirs = [e for e in entries if os.path.isdir(os.path.join(project_output, e))]
    plain_files = [e for e in entries if os.path.isfile(os.path.join(project_output, e))]

    if subdirs:
        print(f"  输出: output/{project_name}/  ({len(subdirs)} 个子目录)")
        for fn in subdirs:
            count = len(os.listdir(os.path.join(project_output, fn)))
            print(f"         └── {fn}/  ({count} 个文件)")
    else:
        print(f"  输出: output/{project_name}/  ({len(plain_files)} 个文件)")

    return True


def deploy_project(project_name, project_config):
    """将生成的文件部署到目标仓库。"""
    deploy_cfg = project_config.get('deploy')
    if not deploy_cfg:
        print(f"  部署跳过: 项目 {project_name} 未配置 deploy")
        print(f"  提示: 请在 config.json 中为 {project_name} 添加 deploy 配置，示例:")
        print(f'    "deploy": {{')
        print(f'        "repo": "~/path/to/your/repo",')
        print(f'        "branch": "target_branch",')
        print(f'        "targets": {{')
        print(f'            "variant_name": "repo_relative/path/to/target/dir"')
        print(f'        }}')
        print(f'    }}')
        return False

    repo_path = os.path.expanduser(deploy_cfg['repo'])
    branch    = deploy_cfg['branch']
    targets   = deploy_cfg['targets']

    print(f"  部署目标: {repo_path}")
    print(f"  目标分支: {branch}")

    # --- 校验仓库 ---
    if not os.path.isdir(repo_path):
        print(f"  ✗ 终止: 仓库路径不存在 - {repo_path}")
        print(f"  提示: 请确认 config.json 中 deploy.repo 路径正确")
        print(f"        支持 ~ 表示用户主目录 (当前展开为 {os.path.expanduser('~')})")
        return False

    ok, _, _ = run_git(repo_path, 'rev-parse', '--git-dir')
    if not ok:
        print(f"  ✗ 终止: 不是有效的 git 仓库 - {repo_path}")
        return False

    # --- 检查并切换分支 ---
    ok, current_branch, _ = run_git(repo_path, 'rev-parse', '--abbrev-ref', 'HEAD')
    if not ok:
        print(f"  ✗ 终止: 无法获取当前分支")
        return False

    if current_branch != branch:
        print(f"  当前分支: {current_branch} → 需要切换到 {branch}")

        ok, status_out, _ = run_git(repo_path, 'status', '--porcelain')
        if not ok:
            print(f"  ✗ 终止: 无法获取仓库状态")
            return False
        if status_out:
            print(f"  ✗ 终止: 仓库有未提交的修改，无法切换分支")
            print(f"  请先手动处理:")
            for line in status_out.split('\n')[:10]:
                print(f"    {line}")
            if status_out.count('\n') > 10:
                print(f"    ... 及更多")
            return False

        ok, _, err = run_git(repo_path, 'checkout', branch)
        if not ok:
            print(f"  ✗ 终止: 切换分支失败 - {err}")
            return False
        print(f"  已切换到分支: {branch}")
    else:
        print(f"  当前分支: {current_branch} (已在目标分支)")

    # --- 拷贝文件 ---
    project_output = os.path.join(OUTPUT_DIR, project_name)
    total_copied = 0
    has_variants = bool(project_config.get('variant_names'))

    for label, target_rel in targets.items():
        if has_variants:
            src_dir = os.path.join(project_output, label)
        else:
            src_dir = project_output

        dst_dir = os.path.join(repo_path, target_rel)

        if not os.path.isdir(src_dir):
            print(f"  警告: 输出目录不存在 - {src_dir}")
            continue

        src_files = [f for f in os.listdir(src_dir) if os.path.isfile(os.path.join(src_dir, f))]
        if not src_files:
            print(f"  警告: {label}/ 下无文件，跳过")
            continue

        os.makedirs(dst_dir, exist_ok=True)

        copied = 0
        for fname in src_files:
            shutil.copy2(os.path.join(src_dir, fname), os.path.join(dst_dir, fname))
            copied += 1

        total_copied += copied
        print(f"  {label}/ → {target_rel}")
        print(f"    {copied} 个文件已拷贝")

    if total_copied > 0:
        print(f"  部署完成: 共 {total_copied} 个文件 (git 提交请手动执行)")
    else:
        print(f"  部署完成: 无文件被拷贝")

    return True


def list_projects(config):
    """列出所有已配置项目及其状态。"""
    projects = config.get('projects', {})
    print(f"已配置项目 ({len(projects)} 个):\n")
    for name, info in projects.items():
        desc = info.get('description', '')
        project_input = os.path.join(INPUT_DIR, name)
        latest = find_latest_excel(project_input) if os.path.isdir(project_input) else None

        print(f"  {name:12s}  {desc}")

        ss_token = info.get('spreadsheet_token', '')
        if ss_token:
            print(f"  {'':12s}  spreadsheet → {ss_token[:12]}... (已配置)")
        else:
            print(f"  {'':12s}  spreadsheet → 未配置 (fetch 不可用)")

        if latest:
            date = extract_date_from_filename(latest)
            print(f"  {'':12s}  input/{name}/  → 最新表格: {os.path.basename(latest)} ({date.strftime('%Y-%m-%d')})")
        else:
            print(f"  {'':12s}  input/{name}/  → 无表格")

        variant_names = info.get('variant_names', {})
        if variant_names:
            print(f"  {'':12s}  变体: {', '.join(variant_names.keys())}")

        deploy_cfg = info.get('deploy')
        if deploy_cfg:
            expanded = os.path.expanduser(deploy_cfg['repo'])
            repo_display = deploy_cfg['repo']
            if expanded != repo_display:
                repo_display = f"{deploy_cfg['repo']}  ({expanded})"
            repo_exists = os.path.isdir(expanded)
            status = "✓" if repo_exists else "✗ 路径不存在"
            print(f"  {'':12s}  deploy → {repo_display}  [{deploy_cfg['branch']}]  {status}")
            for variant, target in deploy_cfg.get('targets', {}).items():
                print(f"  {'':12s}    {variant} → {target}")
        else:
            print(f"  {'':12s}  deploy → 未配置")
        print()


def _resolve_projects(names, all_projects):
    """将命令行中的项目名列表解析为有效项目名。返回列表或 None（有错误时）。"""
    if not names:
        return list(all_projects.keys())
    for p in names:
        if p not in all_projects:
            print(f"错误: 项目 '{p}' 未在 config.json 中配置")
            print(f"可用项目: {', '.join(all_projects.keys())}")
            return None
    return list(names)


def _show_empty_config_hint():
    print("提示: config.json 中尚未配置任何项目。")
    print("请编辑 config.json，添加项目配置。")
    print()
    print("基本配置 (生成文件直接放在 output/项目名/ 下):")
    print('  {')
    print('      "projects": {')
    print('          "项目名": {')
    print('              "description": "项目描述",')
    print('              "excel_pattern": "表格文件名关键词",')
    print('              "spreadsheet_token": "飞书表格 token (可选, fetch 用)"')
    print('          }')
    print('      }')
    print('  }')
    print()
    print("可选字段:")
    print('  "variant_names"       多变体子目录 (如需输出多套文件)')
    print('  "deploy"              部署到 git 仓库 (repo/branch/targets)')
    print('  "spreadsheet_token"   飞书在线表格 token (fetch 自动下载)')
    print()
    print(f"配置文件路径: {CONFIG_FILE}")


KNOWN_ACTIONS = {
    'fetch': 'fetch', 'generate': 'generate', 'gen': 'generate',
    'deploy': 'deploy', 'list': 'list',
    'scan': 'scan', 'snapshot': 'snapshot',
}


def cmd_run(project_names, do_scan, do_fetch, do_generate, do_deploy,
            do_snapshot, all_projects):
    """统一执行流水线：scan → fetch → generate → deploy → snapshot。"""
    target_projects = _resolve_projects(project_names, all_projects)
    if target_projects is None:
        return

    steps = []
    if do_scan:      steps.append('scan')
    if do_fetch:     steps.append('fetch')
    if do_generate:  steps.append('generate')
    if do_deploy:    steps.append('deploy')
    if do_snapshot:  steps.append('snapshot')
    banner = ' + '.join(steps)
    print(f"=== CarPropertyManager v{VERSION}  [{banner}] ===\n")

    # scan / fetch / snapshot 都需要飞书 token
    need_feishu = do_scan or do_fetch or do_snapshot
    feishu_token = None
    if need_feishu:
        feishu_token = get_feishu_token()
        if feishu_token is None:
            return

    # 是否有 scan 之后的后续步骤（决定 scan 是否暂停等待用户确认）
    has_post_scan_steps = do_fetch or do_generate or do_deploy or do_snapshot

    scan_count = 0
    fetch_count = 0
    gen_count = 0
    deploy_count = 0
    snapshot_count = 0

    for project_name in target_projects:
        project_config = all_projects[project_name]
        desc = project_config.get('description', '')
        print(f"[{project_name}] {desc}")

        # --- scan ---
        scan_ok = True
        scan_continue = True  # 用户是否同意继续后续步骤
        if do_scan:
            print()
            print("  --- 扫描 ---")
            result = scan_project(project_name, project_config, feishu_token)
            if result == 'error':
                scan_ok = False
                print(red("  ✘ [scan] 失败"))
            elif result == 'no_diff':
                scan_count += 1
                print(green("  ✓ [scan] 无差异"))
                if has_post_scan_steps:
                    print("  与本地版本一致，跳过后续步骤。")
                    scan_continue = False
            else:  # has_diff
                scan_count += 1
                print(green("  ✓ [scan] 完成"))
                if has_post_scan_steps:
                    try:
                        answer = input("\n  是否继续执行后续步骤？(y/n): ").strip().lower()
                    except (EOFError, KeyboardInterrupt):
                        answer = 'n'
                    if answer not in ('y', 'yes'):
                        scan_continue = False
                        print(yellow("  用户取消，跳过后续步骤。"))

        # --- fetch ---
        fetch_ok = True
        if do_fetch:
            if (not do_scan or scan_ok) and scan_continue:
                print()
                print("  --- 下载 ---")
                if fetch_project(project_name, project_config, feishu_token):
                    fetch_count += 1
                    print(green("  ✓ [fetch] 成功"))
                else:
                    fetch_ok = False
                    print(red("  ✘ [fetch] 失败"))
            else:
                fetch_ok = False
                if not scan_continue:
                    pass  # 用户主动跳过，不再重复提示
                else:
                    print()
                    print(yellow("  ⚠ [fetch] 已跳过（scan 未成功）"))

        # --- generate ---
        gen_ok = True
        if do_generate:
            can_run = scan_continue
            if do_scan and not scan_ok:
                can_run = False
            if do_fetch and not fetch_ok:
                can_run = False
            if can_run:
                print()
                print("  --- 生成 ---")
                if process_project(project_name, project_config):
                    gen_count += 1
                    print(green("  ✓ [generate] 成功"))
                else:
                    gen_ok = False
                    print(red("  ✘ [generate] 失败"))
            else:
                gen_ok = False
                if scan_continue:
                    print()
                    print(yellow("  ⚠ [generate] 已跳过（上游步骤未成功）"))

        # --- deploy ---
        deploy_ok = True
        if do_deploy:
            can_run = scan_continue
            if do_scan and not scan_ok:
                can_run = False
            if do_fetch and not fetch_ok:
                can_run = False
            if do_generate and not gen_ok:
                can_run = False
            if can_run:
                print()
                print("  --- 部署 ---")
                if deploy_project(project_name, project_config):
                    deploy_count += 1
                    print(green("  ✓ [deploy] 成功"))
                else:
                    deploy_ok = False
                    print(red("  ✘ [deploy] 失败"))
            else:
                deploy_ok = False
                if scan_continue:
                    print()
                    print(yellow("  ⚠ [deploy] 已跳过（上游步骤未成功）"))

        # --- snapshot（置于最后，前面全部成功才执行）---
        if do_snapshot:
            can_run = scan_continue
            if do_scan and not scan_ok:
                can_run = False
            if do_fetch and not fetch_ok:
                can_run = False
            if do_generate and not gen_ok:
                can_run = False
            if do_deploy and not deploy_ok:
                can_run = False
            if can_run:
                print()
                print("  --- 版本快照 ---")
                if snapshot_project(project_name, project_config, feishu_token):
                    snapshot_count += 1
                    print(green("  ✓ [snapshot] 成功"))
                else:
                    print(red("  ✘ [snapshot] 失败"))
            else:
                if scan_continue:
                    print()
                    print(yellow("  ⚠ [snapshot] 已跳过（上游步骤未成功）"))

        print()

    # 汇总
    total = len(target_projects)
    parts = []
    if do_scan:
        seg = f"扫描: {scan_count}/{total}"
        parts.append(green(seg) if scan_count == total else red(seg))
    if do_fetch:
        seg = f"下载: {fetch_count}/{total}"
        parts.append(green(seg) if fetch_count == total else red(seg))
    if do_generate:
        seg = f"生成: {gen_count}/{total}"
        parts.append(green(seg) if gen_count == total else red(seg))
    if do_deploy:
        seg = f"部署: {deploy_count}/{total}"
        parts.append(green(seg) if deploy_count == total else red(seg))
    if do_snapshot:
        seg = f"快照: {snapshot_count}/{total}"
        parts.append(green(seg) if snapshot_count == total else red(seg))
    print("  ".join(parts))


def main():
    parser = argparse.ArgumentParser(
        description=f'property — 车辆属性（CarProperty）生成与导入 v{VERSION}',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''操作（可组合，不指定则默认执行完整流水线 scan + fetch + generate + deploy + snapshot）:
  scan               对比飞书在线表格与本地 Excel 差异
  fetch     / -f     从飞书下载表格到 input/
  generate  / -g     从 input/ 读取 Excel 生成代码到 output/
  deploy    / -d     将 output/ 部署到目标 Git 仓库
  snapshot  / -s     为飞书在线表格创建版本快照
  list      / -l     列出所有已配置项目及状态

示例:
  python main.py                        全部项目：完整流水线
  python main.py BAIC                   仅 BAIC：完整流水线
  python main.py scan BAIC              仅 BAIC：仅扫描差异
  python main.py fetch                  全部项目：仅 fetch
  python main.py -f                     同上
  python main.py generate BAIC          仅 BAIC：仅 generate
  python main.py -fg BAIC               仅 BAIC：fetch + generate
  python main.py snapshot BAIC          仅 BAIC：仅创建版本快照
  python main.py list                   列出项目
  python main.py -l                     同上''')

    parser.add_argument('-f', '--fetch', action='store_true',
                        help='从飞书下载表格到 input/')
    parser.add_argument('-g', '--generate', action='store_true',
                        help='从 input/ 读取 Excel 生成代码到 output/')
    parser.add_argument('-d', '--deploy', action='store_true',
                        help='将 output/ 部署到目标 Git 仓库')
    parser.add_argument('-s', '--snapshot', action='store_true',
                        help='为飞书在线表格创建版本快照')
    parser.add_argument('-l', '--list', action='store_true',
                        help='列出所有已配置项目及状态')
    parser.add_argument('-v', '--version', action='version',
                        version=f'property v{VERSION}')
    parser.add_argument('--scan', action='store_true',
                        help='对比飞书在线表格与本地 Excel 差异')
    parser.add_argument('--run', action='store_true', help=argparse.SUPPRESS)
    parser.add_argument('args', nargs='*', metavar='ACTION_OR_PROJECT',
                        help='操作名（scan/fetch/generate/deploy/snapshot/list）或项目名')

    args = parser.parse_args()

    if args.run:
        print("警告: --run 已废弃，请改用 generate 或 -g")
        args.generate = True

    actions = set()
    project_names = []
    for a in args.args:
        key = a.lower()
        if key in KNOWN_ACTIONS:
            actions.add(KNOWN_ACTIONS[key])
        else:
            project_names.append(a)

    if args.fetch:    actions.add('fetch')
    if args.generate: actions.add('generate')
    if args.deploy:   actions.add('deploy')
    if args.snapshot:  actions.add('snapshot')
    if args.list:     actions.add('list')
    if args.scan:     actions.add('scan')

    if not actions:
        actions = {'scan', 'fetch', 'generate', 'deploy', 'snapshot'}

    original_cwd = os.getcwd()
    if _PATHS:
        os.makedirs(str(_PATHS.workspace_dir), exist_ok=True)
        os.chdir(str(_PATHS.workspace_dir))
    else:
        os.chdir(SCRIPT_DIR)

    config = load_config()
    all_projects = config.get('projects', {})

    if not all_projects:
        _show_empty_config_hint()
        os.chdir(original_cwd)
        return

    ensure_project_dirs(all_projects)

    if 'list' in actions:
        list_projects(config)
    else:
        do_scan     = 'scan' in actions
        do_fetch    = 'fetch' in actions
        do_generate = 'generate' in actions
        do_deploy   = 'deploy' in actions
        do_snapshot = 'snapshot' in actions
        cmd_run(project_names, do_scan, do_fetch, do_generate, do_deploy,
                do_snapshot, all_projects)

    os.chdir(original_cwd)


if __name__ == '__main__':
    main()
