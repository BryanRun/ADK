# -*- coding: utf-8 -*-

# @Date    : 2023-11-19
# @Author  : HeZhaoyang,AnChunjie
# @file    : excelop.py
"""excel operation。
   
    excel 表格文件读写模块
    example 1:
        参考《excelop模块API使用说明》
    example 2:
        excel = ExcelOperation(path = './test.xlsx')
        if excel.isopen():
            read_cell = excel['sheeet1'][row][col]
            excel['sheeet1'][row][col] = 'write test'
            excel.close()
"""
import os
from datetime import datetime
import openpyxl
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill, Font, NamedStyle,Border, Side
from enum import Enum
from collections import defaultdict
from itertools import chain




class background_color(Enum):
    WHITE = 'FFFFFF'
    BLACK = '000000'
    RED = 'FF0000'
    GREEN = '00FF00'
    BLUE = '0000FF'
    YELLOW = 'FFFF00'
    CYAN = '00FFFF'
    MAGENTA = 'FF00FF'
    GRAY = '808080'
    SILVER = 'C0C0C0'
    DARK_GRAY = '404040'
    LIGHT_BLUE = 'ADD8E6'
    SKY_BLUE = '87CEEB'
    ORANGE = 'FFA500'
    PURPLE = '800080'
    BROWN = 'A52A2A'
    PINK = 'FFC0CB'
    OLIVE_GREEN = '808000'
    DARK_GREEN = '006400'
    DARK_RED = '8B0000'
    DARK_BLUE = '00008B'
    LIGHT_GRAY = 'D3D3D3'
    PALE_YELLOW = 'FFFFE0'
    LAVENDER = 'E6E6FA'
    TEAL = '4682B4'
    FOREST_GREEN = '228B22'
    CORAL = 'FF7F50'
    LAVENDER_BLUSH = 'FFF0F5'

class font_color(Enum):
    BLACK = '000000'
    RED = 'FF0000'
    GREEN = '00FF00'
    YELLOW = 'FFFF00'
    BLUE = '0000FF'
    DARK_BLUE = '00008B'


class ExcelCell:
    def __init__(self, sheet, row, col) -> None:
        self._sheet = sheet
        self._cell = self._sheet.cell(row=row + 1, column=col + 1)

    def set_value(self, value):
        self._cell.value = value

    def get_value(self):
        return self._cell.value

class ExcelSheetLine:
    def __init__(self, sheet, row) -> None:
        self._sheet = sheet
        self._row = row

    def __getitem__(self, col):
        return ExcelCell(self._sheet, self._row, col)
     
class ExcelSheet:
    def __init__(self, excel: openpyxl.Workbook, name) -> None:
        self._excel = excel
        self._name = name
        if name not in self._excel.sheetnames:
            self._sheet = self._excel.create_sheet(name)
        else:
            self._sheet = self._excel[name]

    def __getitem__(self, row):
        return ExcelSheetLine(self._sheet, row)


class ReadResult:
    def __init__(self,text,first_row,first_col,end_row,end_col) -> None:
        self.text = text
        self.first_row=first_row
        self.end_row =end_row
        self.end_col =end_col


class SheetTemplate():
    def __init__(self,sheet):
        self._sheet = sheet
        pass

    def superaddition(self,write_dict:dict,back_color=background_color.WHITE,\
                        bold=False, size=11,color=font_color.BLACK,real_time_save=False,hotal='left',vtcl='center'):
        if self._sheet.standard == True:
            for col,value in enumerate(self._sheet.header_list):
                col_1 = col+1
                try:
                    self._sheet.write(self._sheet.template_writes_count,col_1,write_dict[str(value)],back_color=back_color,\
                               bold=bold, size=size,color=color,real_time_save=real_time_save,hotal=hotal,vtcl=vtcl)
                except:
                    pass
            self._sheet.template_writes_count += 1
        else:
            for row,value in enumerate(self._sheet.header_list):
                row_1 = row +1
                try:
                    self._sheet.write(row_1,self._sheet.template_writes_count,write_dict[str(value)],back_color=back_color,\
                               bold=bold, size=size,color=color,real_time_save=real_time_save,hotal=hotal,vtcl=vtcl)
                except:
                    pass
            self._sheet.template_writes_count += 1

MAX_CHAR_FIT_COL_WIDTH=120
MAX_LINE_FEED_HEIGHT=5

class Sheet():
    def __init__(self,sheet,work,path) -> None: 
        self._sheet =sheet 
        self._work=work
        self._path=path
        self._col_max_lenght_dict={}
        self._row_max_height_dict={}
        self._col_max_lenght = MAX_CHAR_FIT_COL_WIDTH
        self._row_max_height = MAX_LINE_FEED_HEIGHT

    def find_max_row(self):
        try:
            return  self._sheet.max_row
        except:
            return None

    def find_max_col(self):
        try:
            return  self._sheet.max_column
        except:
            return None

    def write(self,start_row:int,start_column:int,text:str,end_row= None,end_column= None,back_color=background_color.WHITE,\
                    bold=False, size=11,color=font_color.BLACK,auto_column_width=True,real_time_save=False,hotal='left',vtcl='center') -> bool:
        _alignment = Alignment(horizontal=hotal, vertical=vtcl)
        _backgrd_color = PatternFill(start_color=back_color.value,end_color=back_color.value,fill_type='solid')
        _thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        _font_black_bold = Font(color=color.value, bold=bold, size=size)
        if end_row != None or end_column != None:
             self._sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=start_column, end_column=end_column)
        cell=self._sheet.cell(row=start_row,column=start_column)
        cell.value=text
        cell.fill=_backgrd_color
        cell.border=_thin_border
        cell.font=_font_black_bold
        cell.alignment = _alignment
        if auto_column_width == True:
            column_letter = get_column_letter(start_column)
            if not (column_letter in self._col_max_lenght_dict):
                self._col_max_lenght_dict[column_letter] = 0
            if (len(str(cell.value)) >  self._col_max_lenght_dict[column_letter]) and (len(str(cell.value)) < self._col_max_lenght):
                self._col_max_lenght_dict[column_letter] = len(str(cell.value)) 
                self._sheet.column_dimensions[column_letter].width =(self._col_max_lenght_dict[column_letter] * size / 11)+2 
            if not (str(start_row) in self._row_max_height_dict):
                self._row_max_height_dict[str(start_row)] = 0
            line_breaks =len(re.findall(r'\n', cell.value)) 
            if (line_breaks > self._row_max_height_dict[str(start_row)]) and (line_breaks < self._row_max_height):
                self._row_max_height_dict[str(start_row)] = line_breaks 
                self._sheet.row_dimensions[start_row].height = (line_breaks+1)* size * 1.2
        if  real_time_save == True:      
            try:
                self._work.save(self._path)
            except Exception as a :
                return str(a)
        return True

    def read(self,row_num:int,col_num:int,en_mge_cell_chk:bool=True):
        if en_mge_cell_chk == True:
           for merged_cell in self._sheet.merged_cells:
                if (row_num >= merged_cell.min_row and row_num <= merged_cell.max_row) and \
                    (col_num >= merged_cell.min_col and col_num <= merged_cell.max_col):
                    read_text=self._sheet.cell(row=merged_cell.min_row,column=merged_cell.min_col).value
                    return read_text
        _read_text = self._sheet.cell(row=row_num, column=col_num).value 
        return _read_text

    def adjust_column_width(self, target_column=None, font_size=11):
        _font_factor = font_size / 11
        if target_column:
            column_letter = get_column_letter(target_column)
            col = self._sheet[column_letter]
            _max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > _max_length:
                        _max_length = len(cell.value)
                except:
                    pass
        
            _adjusted_width = (_max_length + 2) * 1.0 * _font_factor
            self._sheet.column_dimensions[column_letter].width = _adjusted_width
        else:
            for col in self._sheet.columns:
                _max_length = 0
                column = col[0].column
            
                for cell in col:
                    try:
                        if len(str(cell.value)) > _max_length:
                            _max_length = len(cell.value)
                    except:
                        pass
            
                _adjusted_width = (_max_length + 2) * 1.0 * _font_factor
                self._sheet.column_dimensions[get_column_letter(column)].width = _adjusted_width

    def row_height_simulation_adaptive(self, target_row=None, font_size=11):
        _line_height = font_size * 1.2

        _rows_to_adjust = [target_row] if target_row else range(1, self._sheet.max_row + 1)
    
        for row_idx in _rows_to_adjust:
            row = self._sheet[row_idx]
            lines = max(len(re.findall(r'\n', str(cell.value))) + 1 for cell in row if cell.value is not None)
            _adjusted_height = lines * _line_height
            self._sheet.row_dimensions[row_idx].height = _adjusted_height

    def create_template(self,header:list,header_color=background_color.LIGHT_BLUE,standard:bool=True,bold=True,\
                         size=14,color=font_color.BLACK,hotal='center',vtcl='center'):
        self.header_list = header
        self.standard= standard
        if standard == True:
            for col,value in enumerate(header):
                col_1 = col+1
                self.write(1,col_1,str(value),back_color=header_color,bold=bold,size=size,color=color,hotal=hotal,vtcl=vtcl)
        else:
            for line,value in enumerate(header):
                line_1=line+1
                self.write(line_1,1,str(value),back_color=header_color,bold=bold,size=size,color=color,hotal=hotal,vtcl=vtcl)
        self.template_writes_count = 2
        return SheetTemplate(self)


class ExcelOperation:
    def __init__(self, path:str,new_build:bool=False,sheetname:str="autosheet") -> None:
        self._is_open = False
        self._path: str =os.path.normpath(path).replace("\\", "/")
        if os.path.dirname(self._path):
            try:
                self._excel = load_workbook(self._path)
                self._is_open = True
            except Exception as a :
                if new_build == True: 
                    try:
                        if self._path.split('/')[-1].strip()[-4:] == "xlsx":
                            try:
                                os.remove(self._path)
                            except:
                                print("Excel remove fail!!!") 
                            workbook = openpyxl.Workbook()
                            sheet=workbook.active
                            sheet.title=sheetname
                            workbook.save(self._path)
                            self._excel = load_workbook(self._path)
                            self._is_open = True
                        else:
                            self._is_open = False
                            raise ValueError("workbook open fail:No xlsx file name!")
                    except PermissionError as e: 
                        self._is_open = False
                        raise ValueError("workbook open fail:"+str(e))
                else:
                    self._is_open = False
                    raise ValueError("workbook open fail:"+str(a))
        else:
            self._is_open = False
            raise ValueError("workbook open fail:routing problem")
        
    def isopen(self) -> bool:
        return self._is_open

    def create_sheet(self, sheet_name:str,index:int=1) -> str:
        if self._is_open == True:
            sheet_title=self._excel.create_sheet(sheet_name, index)
            return sheet_title.title
        else:
            return None

    def delete_sheet(self, sheet_name:str) -> bool:
        if self._is_open == True:
            self._excel.remove(self._excel[sheet_name])
            return True
        else:
            return False
        
    def open_sheet(self,sheet_name):
        if self._is_open == True:
            try:
                if isinstance(sheet_name, str):
                    return Sheet( self._excel[sheet_name],self._excel,self._path)
                elif isinstance(sheet_name, int):
                    return Sheet(self._excel.worksheets[sheet_name],self._excel,self._path)
                else:
                    return ("The parameter passing type is incorrect!")
            except Exception as a :
                return str(a)
        else:
            return ("excel open fail!")

    def get_sheet(self,sheet_name:str) -> any:
        if self._is_open == True:
            for sheet in self._excel.sheetnames:
                if sheet == sheet_name:
                    return  self._excel[sheet_name]
        else:
            return None

    def find_sheet(self,sheet_name:str) -> bool:
        if self._is_open == True:
            for sheetna in self._excel.sheetnames:
                if sheetna == sheet_name:
                    return True
        else:
            pass
        return False

    def get_sheet_name_list(self):
        if self._is_open == True: 
            return self._excel.sheetnames
        else:
            return None

    def save(self):
        try:
            self._excel.save(self._path)
        except Exception as a :
            return str(a)

    def close(self):
        try:
            self._excel.save(self._path)
            del  self._excel
        except Exception as a :
            return str(a)

    def __getitem__(self, name):
        if isinstance(name, str):
            return ExcelSheet(self._excel, name)
        else:
            return None
